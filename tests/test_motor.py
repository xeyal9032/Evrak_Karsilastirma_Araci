# -*- coding: utf-8 -*-
"""karsilastir_motor birim testleri."""
import os
import tempfile
import unittest
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
import sys
sys.path.insert(0, ROOT)

import karsilastir_motor as motor

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
SAMPLE_A = os.path.join(FIXTURES, "sample_a.csv")
SAMPLE_B = os.path.join(FIXTURES, "sample_b.csv")
SAMPLE_JOURNAL = os.path.join(FIXTURES, "sample_journal.xlsx")
SAMPLE_DATEV_XLSX = os.path.join(FIXTURES, "sample_datev.xlsx")
MART_DIR = r"C:\Users\xeyal\Desktop\excel1\mart"


class TestLoad(unittest.TestCase):
    def test_load_sample_a(self):
        entries = motor.load(SAMPLE_A)
        self.assertEqual(len(entries), 5)
        self.assertEqual(entries[0]["belegfeld1"], "RE-1001")
        self.assertEqual(entries[0]["amt"], 1000.0)

    def test_format_error_missing_column(self):
        with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="latin-1", newline="") as f:
            f.write("meta\n")
            f.write("Umsatz;Konto\n")
            f.write("1,00;1000\n")
            path = f.name
        try:
            with self.assertRaises(motor.FormatError):
                motor.load(path)
        finally:
            os.unlink(path)

    def test_detect_datev_csv(self):
        self.assertEqual(motor.detect_format(SAMPLE_A), motor.FORMAT_DATEV_CSV)

    def test_detect_journal_xlsx(self):
        self.assertEqual(motor.detect_format(SAMPLE_JOURNAL), motor.FORMAT_JOURNAL_XLSX)

    def test_detect_datev_xlsx(self):
        self.assertEqual(motor.detect_format(SAMPLE_DATEV_XLSX), motor.FORMAT_DATEV_XLSX)

    def test_load_journal(self):
        entries = motor.load(SAMPLE_JOURNAL)
        self.assertEqual(len(entries), 3)
        self.assertEqual(entries[0]["belegfeld1"], "RE-1001")
        self.assertEqual(entries[0]["konto"], "1200")
        self.assertEqual(entries[0]["gegenkonto"], "8400")
        self.assertEqual(entries[0]["datum"], "1501")
        self.assertEqual(entries[0]["sh"], "S")

    def test_load_datev_xlsx(self):
        entries = motor.load(SAMPLE_DATEV_XLSX)
        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0]["belegfeld1"], "RE-1001")
        self.assertEqual(entries[0]["amt"], 1000.0)


class TestCompare(unittest.TestCase):
    def test_exact_match_and_mismatch_and_only(self):
        e1, e2, net1, net2 = motor.compare(SAMPLE_A, SAMPLE_B)

        self.assertEqual(net1, 1)  # storno + original netted
        self.assertEqual(net2, 0)

        by_beleg1 = {e["belegfeld1"]: e for e in e1 if not e["netted"]}
        self.assertTrue(by_beleg1["RE-1001"]["matched"])
        self.assertEqual(by_beleg1["RE-1001"]["pair_kind"], "MATCH")

        self.assertTrue(by_beleg1["RE-2002"]["matched"])
        self.assertEqual(by_beleg1["RE-2002"]["pair_kind"], "MISMATCH")

        self.assertFalse(by_beleg1["RE-ONLY1"]["matched"])

        only2 = [e for e in e2 if e["belegfeld1"] == "RE-ONLY2"][0]
        self.assertFalse(only2["matched"])

    def test_datev_csv_vs_journal_xlsx(self):
        e1, e2, net1, net2 = motor.compare(SAMPLE_A, SAMPLE_JOURNAL)
        by_beleg1 = {e["belegfeld1"]: e for e in e1 if not e["netted"]}
        self.assertTrue(by_beleg1["RE-1001"]["matched"])
        self.assertEqual(by_beleg1["RE-1001"]["pair_kind"], "MATCH")
        self.assertTrue(by_beleg1["RE-2002"]["matched"])
        self.assertEqual(by_beleg1["RE-2002"]["pair_kind"], "MISMATCH")
        self.assertFalse(by_beleg1["RE-ONLY1"]["matched"])

    def test_build_report_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "report.xlsx")
            result = motor.build_report(SAMPLE_A, SAMPLE_B, out, f1_label="A", f2_label="B")
            self.assertTrue(os.path.isfile(out))
            self.assertEqual(result["match_count"], 1)
            self.assertEqual(result["mismatch_count"], 1)
            self.assertEqual(result["only1_count"], 1)
            self.assertEqual(result["only2_count"], 1)

            wb = openpyxl.load_workbook(out)
            self.assertIn("KARSILASTIRMA-SRAVNENIE", wb.sheetnames)
            self.assertIn("OZET-SVODKA", wb.sheetnames)
            self.assertIn("FARKLAR-RAZNICA", wb.sheetnames)
            self.assertIn("HESAP FARKI-SCHETA", wb.sheetnames)

    def test_build_report_csv_vs_journal(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "report_journal.xlsx")
            result = motor.build_report(
                SAMPLE_A, SAMPLE_JOURNAL, out,
                f1_label="DATEV", f2_label="Journal",
            )
            self.assertTrue(os.path.isfile(out))
            self.assertGreaterEqual(result["match_count"], 1)
            self.assertGreaterEqual(result["mismatch_count"], 1)


@unittest.skipUnless(os.path.isdir(MART_DIR), "mart klasoru yok")
class TestMartIntegration(unittest.TestCase):
    def test_mart_datev_vs_journal(self):
        files = os.listdir(MART_DIR)
        csvs = [f for f in files if f.lower().endswith(".csv")]
        xlsxs = [f for f in files if f.lower().endswith(".xlsx")]
        self.assertTrue(csvs and xlsxs)
        csv_path = os.path.join(MART_DIR, csvs[0])
        xlsx_path = os.path.join(MART_DIR, xlsxs[0])
        self.assertEqual(motor.detect_format(csv_path), motor.FORMAT_DATEV_CSV)
        self.assertEqual(motor.detect_format(xlsx_path), motor.FORMAT_JOURNAL_XLSX)
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "mart_report.xlsx")
            result = motor.build_report(
                csv_path, xlsx_path, out,
                f1_label="DATEV", f2_label="Journal",
            )
            # Onceki manuel test bandi: ~649 eslesme
            self.assertGreaterEqual(result["match_count"], 600)
            self.assertEqual(result["f1_total"], 689)
            self.assertEqual(result["f2_total"], 812)


class TestHelpers(unittest.TestCase):
    def test_norm_amount(self):
        self.assertEqual(motor.norm_amount("1.234,56"), 1234.56)
        self.assertEqual(motor.norm_amount(1000.5), 1000.5)
        self.assertIsNone(motor.norm_amount(""))

    def test_norm_belegfeld(self):
        self.assertEqual(motor.norm_belegfeld("RE 2026.01"), "RE202601")
        self.assertEqual(motor.norm_belegfeld("re 1.2,3"), "RE123")

    def test_date_sort_key(self):
        self.assertEqual(motor.date_sort_key("1501"), (1, 15))
        self.assertEqual(motor.date_sort_key("bad"), (99, 99))

    def test_format_belegdatum(self):
        self.assertEqual(motor.format_belegdatum(datetime(2026, 3, 1)), "0103")
        self.assertEqual(motor.format_belegdatum("01.03.2026"), "0103")


if __name__ == "__main__":
    unittest.main()
