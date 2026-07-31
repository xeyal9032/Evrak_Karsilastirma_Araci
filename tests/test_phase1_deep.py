# -*- coding: utf-8 -*-
"""
Faz 1 / motor derin testleri:
- Durum sutunu ↔ grup kind birebir
- FILTRE sayfa satir sayilari ↔ sonuc sayaclari
- Progress stage zinciri
- Coklu bos sheet / tum sheet bos
- CLI format hatasi, quiet/non-quiet, lang
- app_log dosya yazimi
- i18n fallback
- GUI dil degisimi smoke
"""
import logging
import os
import sys
import tempfile
import unittest
from collections import Counter
from datetime import datetime
from io import StringIO

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import app_log
import i18n
import karsilastir_motor as motor
from karsilastir import App, _wants_cli, run_cli
from tests.helpers_datev import DATEV_HEADER_LONG, drow, write_datev_csv, write_datev_xlsx


def _tk_display_available():
    import tkinter as tk

    try:
        root = tk.Tk()
        root.destroy()
        return True
    except tk.TclError:
        return False


def _rgb(cell):
    fill = cell.fill
    if not fill or not fill.fgColor or not fill.fgColor.rgb:
        return None
    r = str(fill.fgColor.rgb).upper()
    return r[-6:] if len(r) >= 6 else r


def _reset_app_log():
    """Test izolasyonu icin logger singleton'ini sifirla."""
    log = logging.getLogger("evrak_karsilastirma")
    for h in list(log.handlers):
        try:
            h.close()
        except Exception:
            pass
        log.removeHandler(h)
    app_log._LOGGER = None


class TestDurumColumnDeep(unittest.TestCase):
    def test_every_combined_row_durum_matches_group_kind(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("20,00", "S", "1", "2", "0202", "D", "diff"),
                drow("30,00", "S", "3", "4", "0303", "O1", "only1"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("21,00", "S", "1", "2", "0202", "D", "diff"),
                drow("40,00", "S", "5", "6", "0404", "O2", "only2"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="L1", f2_label="L2")
            e1, e2, *_ = motor.compare(f1, f2)
            groups = motor.build_combined_rows(e1, e2)

            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 2 * len(groups))
            for gi, g in enumerate(groups):
                r1, r2 = rows[gi * 2], rows[gi * 2 + 1]
                self.assertEqual(r1[-1], g["kind"])
                self.assertEqual(r2[-1], g["kind"])
                # Not sutunu hala index 9; Durum index 10
                self.assertEqual(len(r1), 11)

    def test_durum_counter_matches_result_dict(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("20,00", "S", "1", "2", "0202", "D", "a"),
                drow("30,00", "S", "1", "2", "0303", "X", "only"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("99,00", "S", "1", "2", "0202", "D", "a"),
                drow("40,00", "S", "1", "2", "0404", "Y", "only2"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            result = motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            # Her grup 2 satir -> Durum sayaci / 2 == group count
            c = Counter()
            for row in ws.iter_rows(min_row=2, values_only=True):
                c[row[-1]] += 1
            self.assertEqual(c["MATCH"] // 2, result["match_count"])
            self.assertEqual(c["MISMATCH"] // 2, result["mismatch_count"])
            self.assertEqual(c["ONLY1"] // 2, result["only1_count"])
            self.assertEqual(c["ONLY2"] // 2, result["only2_count"])


class TestFilterSheetsDeep(unittest.TestCase):
    def test_filtre_row_counts_exact(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("20,00", "S", "1", "2", "0202", "D1", "a"),
                drow("25,00", "S", "1", "2", "0203", "D2", "b"),
                drow("30,00", "S", "1", "2", "0303", "O1", "x"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("21,00", "S", "1", "2", "0202", "D1", "a"),
                drow("26,00", "S", "1", "2", "0203", "D2", "b"),
                drow("40,00", "S", "1", "2", "0404", "O2", "y"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)

            red_n = sum(1 for _ in wb["FILTRE-KIRMIZI"].iter_rows(min_row=2, values_only=True))
            orange_n = sum(1 for _ in wb["FILTRE-TURUNCU"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(red_n, 2 * (r["only1_count"] + r["only2_count"]))
            self.assertEqual(orange_n, 2 * r["mismatch_count"])
            self.assertGreaterEqual(r["mismatch_count"], 2)

            # Kirmizi sayfada MATCH olmamali
            for row in wb["FILTRE-KIRMIZI"].iter_rows(min_row=2, values_only=True):
                self.assertIn(row[-1], ("ONLY1", "ONLY2"))
                self.assertNotEqual(row[-1], "MATCH")

            # Turuncu renk dogrulama
            for row in wb["FILTRE-TURUNCU"].iter_rows(min_row=2):
                self.assertEqual(row[-1].value, "MISMATCH")
                self.assertEqual(_rgb(row[0]), "FFC000")

    def test_perfect_match_filtre_sheets_header_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "B", "t")])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out)
            self.assertEqual(r["match_count"], 1)
            self.assertEqual(r["mismatch_count"], 0)
            self.assertEqual(r["only1_count"] + r["only2_count"], 0)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(sum(1 for _ in wb["FILTRE-KIRMIZI"].iter_rows(min_row=2)), 0)
            self.assertEqual(sum(1 for _ in wb["FILTRE-TURUNCU"].iter_rows(min_row=2)), 0)
            # Bos filtre sayfasinda auto_filter opsiyonel; header yine var
            self.assertEqual(wb["FILTRE-KIRMIZI"].cell(1, 11).value, "Durum / Status")

    def test_filtre_subset_of_combined_belegs(self):
        """Filtre sayfalarindaki Belegfeld1 degerleri combined sheet'te ayni Durum ile var."""
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "SAME", "ok"),
                drow("20,00", "S", "1", "2", "0202", "DIFF", "a"),
                drow("30,00", "S", "1", "2", "0303", "ONLYA", "x"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "SAME", "ok"),
                drow("99,00", "S", "1", "2", "0202", "DIFF", "a"),
                drow("40,00", "S", "1", "2", "0404", "ONLYB", "y"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            combined = wb["KARSILASTIRMA-SRAVNENIE"]
            by_status = {}
            for row in combined.iter_rows(min_row=2, values_only=True):
                status, beleg = row[-1], str(row[7] or "")
                by_status.setdefault(status, set()).add(beleg)

            for row in wb["FILTRE-TURUNCU"].iter_rows(min_row=2, values_only=True):
                beleg = str(row[7] or "")
                if beleg:
                    self.assertIn(beleg, by_status.get("MISMATCH", set()))

            for row in wb["FILTRE-KIRMIZI"].iter_rows(min_row=2, values_only=True):
                beleg = str(row[7] or "")
                status = row[-1]
                if beleg:
                    self.assertIn(beleg, by_status.get(status, set()))


class TestProgressDeep(unittest.TestCase):
    EXPECTED_STAGES = [
        "load1", "load2", "netting", "match_tier1", "match_tier2", "match_tier3",
        "match_value", "match_cross", "match_tier4", "compare_done",
        "build_sheets", "done",
    ]

    def test_stage_sequence_and_bounds(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "B", "t")])
            events = []
            motor.build_report(
                f1, f2, os.path.join(tmp, "o.xlsx"),
                progress_cb=lambda p, s: events.append((p, s)),
            )
            stages = [s for _, s in events]
            for s in self.EXPECTED_STAGES:
                self.assertIn(s, stages, msg=f"missing stage {s}")
            # siralama: her expected stage bir oncekilerden sonra gelmeli
            idxs = [stages.index(s) for s in self.EXPECTED_STAGES]
            self.assertEqual(idxs, sorted(idxs))
            for pct, _ in events:
                self.assertGreaterEqual(pct, 0)
                self.assertLessEqual(pct, 100)
            self.assertEqual(events[0][0], 5)
            self.assertEqual(events[-1], (100, "done"))

    def test_compare_alone_stops_at_compare_done(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            events = []
            motor.compare(f1, f2, progress_cb=lambda p, s: events.append((p, s)))
            self.assertEqual(events[-1][1], "compare_done")
            self.assertNotIn("done", [s for _, s in events])
            self.assertNotIn("build_sheets", [s for _, s in events])

    def test_none_progress_cb_no_crash(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            motor.build_report(f1, f2, os.path.join(tmp, "o.xlsx"), progress_cb=None)


class TestEmptySheetsDeep(unittest.TestCase):
    def test_two_empty_journal_sheets_then_data(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "j.xlsx")
            wb = openpyxl.Workbook()
            hdr = ["Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                   "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto"]
            wb.active.title = "E1"
            wb.active.append(hdr)
            ws2 = wb.create_sheet("E2")
            ws2.append(["title"])
            ws2.append(hdr)
            ws3 = wb.create_sheet("OK")
            ws3.append(["Journal"])
            ws3.append(hdr)
            ws3.append([datetime(2026, 3, 1), None, 1, "Z9", "ok", 50.0, "EUR", "10", "20", None, None])
            wb.save(path)
            wb.close()
            entries = motor.load_journal_xlsx(path)
            self.assertEqual(len(entries), 1)
            self.assertEqual(entries[0]["belegfeld1"], "Z9")

    def test_all_journal_sheets_empty_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "j.xlsx")
            wb = openpyxl.Workbook()
            hdr = ["Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                   "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto"]
            wb.active.append(hdr)
            wb.create_sheet("E2").append(hdr)
            wb.save(path)
            wb.close()
            with self.assertRaises(motor.FormatError):
                motor.load_journal_xlsx(path)

    def test_all_datev_sheets_empty_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "d.xlsx")
            wb = openpyxl.Workbook()
            wb.active.append(["EXTF"])
            wb.active.append(DATEV_HEADER_LONG)
            s2 = wb.create_sheet("E2")
            s2.append(["EXTF"])
            s2.append(DATEV_HEADER_LONG)
            wb.save(path)
            wb.close()
            with self.assertRaises(motor.FormatError):
                motor.load_datev_xlsx(path)

    def test_detect_format_uses_second_sheet_journal(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "j.xlsx")
            wb = openpyxl.Workbook()
            wb.active.title = "Noise"
            wb.active.append(["foo", "bar"])
            ws = wb.create_sheet("J")
            ws.append(["Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                       "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto"])
            ws.append([datetime(2026, 1, 1), None, 1, "A", "t", 1.0, "EUR", "1", "2", None, None])
            wb.save(path)
            wb.close()
            self.assertEqual(motor.detect_format(path), motor.FORMAT_JOURNAL_XLSX)


class TestCLIDeep(unittest.TestCase):
    def test_cli_bad_format_exit_1(self):
        with tempfile.TemporaryDirectory() as tmp:
            bad = os.path.join(tmp, "bad.csv")
            with open(bad, "w", encoding="utf-8") as f:
                f.write("not;a;datev;file\n")
            good = os.path.join(tmp, "good.csv")
            write_datev_csv(good, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([bad, good, "--quiet"])
            finally:
                sys.stdout, sys.stderr = old
            self.assertEqual(code, 1)

    def test_cli_progress_on_stderr_when_not_quiet(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            out = os.path.join(tmp, "o.xlsx")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([f1, f2, "-o", out, "--lang", "de"])
                err = sys.stderr.getvalue()
                out_txt = sys.stdout.getvalue()
            finally:
                sys.stdout, sys.stderr = old
            self.assertEqual(code, 0)
            self.assertIn("100%", err)
            self.assertIn("done", err)
            self.assertIn("OK", out_txt)
            self.assertIn("match=", out_txt)

    def test_cli_lang_de_ok_message(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            out = os.path.join(tmp, "o.xlsx")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([f1, f2, "-o", out, "--lang", "de", "--quiet"])
                txt = sys.stdout.getvalue()
            finally:
                sys.stdout, sys.stderr = old
            self.assertEqual(code, 0)
            self.assertIn("OK ->", txt)

    def test_wants_cli_helpers(self):
        self.assertFalse(_wants_cli([]))
        self.assertTrue(_wants_cli(["--help"]))
        self.assertTrue(_wants_cli(["a.csv", "b.csv"]))


class TestAppLogDeep(unittest.TestCase):
    def test_cli_writes_log_lines(self):
        _reset_app_log()
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            out = os.path.join(tmp, "o.xlsx")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                self.assertEqual(run_cli([f1, f2, "-o", out, "--quiet"]), 0)
            finally:
                sys.stdout, sys.stderr = old
            # Flush handlers
            log = app_log.get_logger()
            for h in log.handlers:
                h.flush()
            day = __import__("datetime").datetime.now().strftime("%Y%m%d")
            path = os.path.join(app_log.LOG_DIR, f"compare_{day}.log")
            self.assertTrue(os.path.isfile(path))
            with open(path, encoding="utf-8") as fh:
                text = fh.read()
            self.assertIn("compare start (cli)", text)
            self.assertIn("ok out=", text)
            self.assertIn(os.path.basename(f1), text)
        _reset_app_log()

    def test_log_run_err_records_traceback(self):
        _reset_app_log()
        app_log.log_run_err(ValueError("boom"), "TRACE_LINE_XYZ")
        log = app_log.get_logger()
        for h in log.handlers:
            h.flush()
        day = __import__("datetime").datetime.now().strftime("%Y%m%d")
        path = os.path.join(app_log.LOG_DIR, f"compare_{day}.log")
        with open(path, encoding="utf-8") as fh:
            text = fh.read()
        self.assertIn("boom", text)
        self.assertIn("TRACE_LINE_XYZ", text)
        _reset_app_log()


class TestI18nDeep(unittest.TestCase):
    def test_unknown_lang_falls_back_to_tr(self):
        self.assertEqual(i18n.t("compare", lang="xx"), i18n.t("compare", lang="tr"))

    def test_missing_key_returns_key(self):
        self.assertEqual(i18n.t("___no_such_key___", lang="en"), "___no_such_key___")

    def test_all_langs_nonempty_values(self):
        for lang in i18n.SUPPORTED:
            for key in i18n.locale_keys(lang):
                val = i18n.t(key, lang=lang)
                self.assertTrue(str(val).strip(), msg=f"{lang}:{key}")


@unittest.skipUnless(_tk_display_available(), "Tk display unavailable")
class TestGuiLangDeep(unittest.TestCase):
    def test_app_language_switch_updates_labels(self):
        app = App()
        try:
            app.lang.set("en")
            app._apply_lang()
            self.assertEqual(app.compare_btn.cget("text"), "Compare")
            app.lang.set("de")
            app._apply_lang()
            self.assertEqual(app.compare_btn.cget("text"), "Vergleichen")
            app.lang.set("ru")
            app._apply_lang()
            self.assertIn("Сравнить", app.compare_btn.cget("text"))
        finally:
            app.destroy()

    def test_assign_path_rejects_bad_format(self):
        import tkinter.messagebox as mb
        app = App()
        shown = []

        def fake_error(title, msg):
            shown.append((title, msg))

        old = mb.showerror
        mb.showerror = fake_error
        try:
            with tempfile.TemporaryDirectory() as tmp:
                bad = os.path.join(tmp, "x.txt")
                with open(bad, "w", encoding="utf-8") as f:
                    f.write("nope")
                app._assign_path(1, bad)
                self.assertEqual(app.file1.get(), "")
                self.assertTrue(shown)
        finally:
            mb.showerror = old
            app.destroy()

    def test_assign_path_accepts_valid_csv(self):
        import tkinter.messagebox as mb
        app = App()
        mb.showerror = lambda *a, **k: None
        try:
            with tempfile.TemporaryDirectory() as tmp:
                f1 = os.path.join(tmp, "a.csv")
                write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "B", "t")])
                app._assign_path(1, f1)
                self.assertEqual(app.file1.get(), f1)
                self.assertEqual(app._fmt1, motor.FORMAT_DATEV_CSV)
        finally:
            app.destroy()


class TestAutofilterDeep(unittest.TestCase):
    def test_combined_autofilter_covers_all_data_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            rows = [drow(f"{i},00", "S", "1", "2", "0101", f"B{i}", f"t{i}") for i in range(1, 6)]
            write_datev_csv(f1, rows)
            write_datev_csv(f2, rows[:3] + [drow("99,00", "S", "1", "2", "0101", "B4", "t4"),
                                            drow("5,00", "S", "9", "8", "0505", "ONLY2", "z")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            n = sum(1 for _ in ws.iter_rows(min_row=2))
            self.assertTrue(ws.auto_filter.ref.endswith(str(n + 1)))
            self.assertTrue(ws.auto_filter.ref.startswith("A1:K"))


class TestWriteFilterSheetDirect(unittest.TestCase):
    def test_write_filter_sheet_kinds_param(self):
        e1 = [motor.make_entry(1, "10,00", "S", "1", "2", "0101", "A", "t", 10.0)]
        e2 = [motor.make_entry(1, "11,00", "S", "1", "2", "0101", "A", "t", 11.0)]
        e1[0]["matched"] = True
        e1[0]["pair"] = e2[0]
        e1[0]["pair_kind"] = "MISMATCH"
        e2[0]["matched"] = True
        e2[0]["pair"] = e1[0]
        e2[0]["pair_kind"] = "MISMATCH"
        groups = motor.build_combined_rows(e1, e2)
        wb = openpyxl.Workbook()
        ws = wb.active
        motor.write_filter_sheet(ws, groups, "A", "B", {"MISMATCH"})
        data = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data), 2)
        self.assertEqual(data[0][-1], "MISMATCH")
        motor.write_filter_sheet(wb.create_sheet("x"), groups, "A", "B", {"ONLY1"})
        self.assertEqual(sum(1 for _ in wb["x"].iter_rows(min_row=2)), 0)


if __name__ == "__main__":
    unittest.main()
