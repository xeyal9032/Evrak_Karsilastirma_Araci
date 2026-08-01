# -*- coding: utf-8 -*-
"""Daha fazla e2e: yeniden uretim kararliligi, FARKLAR kaynak etiketi, buyuk rastgele."""
import os
import random
import sys
import tempfile
import unittest

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(__file__))

import karsilastir_motor as motor
from fixture_paths import mart_dir, excel1_dir, has_mart
from helpers_datev import drow, write_datev_csv

MART = mart_dir() or ""


class TestE2EFarklarSourceLabels(unittest.TestCase):
    def test_farklar_source_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "ONLYA", "a"),
                drow("20,00", "S", "1", "2", "0202", "BOTH", "b"),
            ])
            write_datev_csv(f2, [
                drow("20,00", "S", "1", "2", "0202", "BOTH", "b"),
                drow("30,00", "S", "3", "4", "0303", "ONLYB", "c"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="Alpha", f2_label="Beta")
            wb = openpyxl.load_workbook(out, data_only=True)
            ws = wb["FARKLAR-RAZNICA"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            sources = {r[0] for r in rows}
            belegs = {r[7] for r in rows}
            self.assertEqual(sources, {"Alpha", "Beta"})
            self.assertEqual(belegs, {"ONLYA", "ONLYB"})
            self.assertNotIn("BOTH", belegs)


class TestE2ERandomStableSeed(unittest.TestCase):
    def test_seeded_random_dataset_deterministic(self):
        def build_rows(seed, n=150, overlap=120):
            rng = random.Random(seed)
            shared = []
            for i in range(overlap):
                shared.append(drow(
                    f"{rng.randint(1, 999)},{rng.randint(0, 99):02d}",
                    "S",
                    str(rng.randint(1000, 1999)),
                    str(rng.randint(2000, 2999)),
                    f"{rng.randint(1, 28):02d}{rng.randint(1, 12):02d}",
                    f"S{i:04d}",
                    f"Shared {i}",
                ))
            only1 = [
                drow("1,00", "S", "9001", "9002", "0101", f"O1{i}", f"Only1 {i}")
                for i in range(n - overlap)
            ]
            only2 = [
                drow("2,00", "S", "9101", "9102", "0202", f"O2{i}", f"Only2 {i}")
                for i in range(n - overlap)
            ]
            return shared + only1, shared + only2

        rows1, rows2 = build_rows(42)
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, rows1)
            write_datev_csv(f2, rows2)
            r1 = motor.build_report(f1, f2, os.path.join(tmp, "1.xlsx"), f1_label="A", f2_label="B")
            r2 = motor.build_report(f1, f2, os.path.join(tmp, "2.xlsx"), f1_label="A", f2_label="B")
        self.assertEqual(r1["match_count"], 120)
        self.assertEqual(r1["only1_count"], 30)
        self.assertEqual(r1["only2_count"], 30)
        for k in ("match_count", "mismatch_count", "only1_count", "only2_count", "f1_total", "f2_total"):
            self.assertEqual(r1[k], r2[k])


class TestE2EMartTripleRun(unittest.TestCase):
    @unittest.skipUnless(has_mart(), "mart yok")
    def test_three_consecutive_runs_identical_counts(self):
        files = os.listdir(MART)
        cpath = os.path.join(MART, [f for f in files if f.endswith(".csv")][0])
        xpath = os.path.join(MART, [f for f in files if f.endswith(".xlsx")][0])
        counts = []
        with tempfile.TemporaryDirectory() as tmp:
            for i in range(3):
                r = motor.build_report(
                    cpath, xpath, os.path.join(tmp, f"{i}.xlsx"),
                    f1_label="D", f2_label="J",
                )
                counts.append((
                    r["match_count"], r["mismatch_count"],
                    r["only1_count"], r["only2_count"],
                    r["f1_total"], r["f2_total"],
                ))
        self.assertEqual(counts[0], counts[1])
        self.assertEqual(counts[1], counts[2])
        self.assertEqual(counts[0], (644, 6, 39, 162, 689, 812))


class TestE2EEntrySheetStatusColumn(unittest.TestCase):
    def test_unmatched_status_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "A", "only")])
            write_datev_csv(f2, [drow("2,00", "S", "3", "4", "0202", "B", "other")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="Left", f2_label="Right")
            wb = openpyxl.load_workbook(out, data_only=True)
            status = wb["Left"].cell(2, 9).value
            self.assertIn("SADECE", status.upper())
            self.assertIn("DOSYADA", status.upper())


class TestE2ECombinedSourceLabels(unittest.TestCase):
    def test_source_column_alternates(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "M", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "M", "t")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="DosyaBir", f2_label="DosyaIki")
            wb = openpyxl.load_workbook(out, data_only=True)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            sources = [row[0] for row in ws.iter_rows(min_row=2, max_row=3, values_only=True)]
            self.assertEqual(sources, ["DosyaBir", "DosyaIki"])


if __name__ == "__main__":
    unittest.main()
