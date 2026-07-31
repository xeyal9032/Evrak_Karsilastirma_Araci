# -*- coding: utf-8 -*-
"""
E2E testler: load -> compare -> build_report -> Excel icerik dogrulama.
Gercek mart klasoru + sentetik dosyalar. Harf/rakam bozulmadan raporlanmali.
"""
import csv
import os
import sys
import tempfile
import unittest
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
import karsilastir_motor as motor

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
SAMPLE_A = os.path.join(FIXTURES, "sample_a.csv")
SAMPLE_B = os.path.join(FIXTURES, "sample_b.csv")
SAMPLE_JOURNAL = os.path.join(FIXTURES, "sample_journal.xlsx")
SAMPLE_DATEV_XLSX = os.path.join(FIXTURES, "sample_datev.xlsx")
MART_DIR = r"C:\Users\xeyal\Desktop\excel1\mart"
EXCEL1 = r"C:\Users\xeyal\Desktop\excel1"

DATEV_HEADER = [
    "Umsatz", "Soll/Haben-Kennzeichen", "WKZ Umsatz", "Kurs", "Basisumsatz", "WKZ Basisumsatz",
    "Konto", "Gegenkonto", "BU-Schluessel", "Belegdatum", "Belegfeld 1", "Belegfeld 2",
    "Skonto", "Buchungstext",
]


def _write_datev(path, data_rows):
    with open(path, "w", encoding="latin-1", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["EXTF", "700", "21", "Buchungsstapel", "e2e"])
        w.writerow(DATEV_HEADER)
        for r in data_rows:
            w.writerow(r)


def _row(u, sh, k, g, d, b, t):
    return [u, sh, "", "", "", "", k, g, "", d, b, "", "", t]


class TestE2EReportStructure(unittest.TestCase):
    def test_report_has_all_sheets_and_legend(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            result = motor.build_report(SAMPLE_A, SAMPLE_B, out, f1_label="A", f2_label="B")
            self.assertTrue(os.path.isfile(out))
            wb = openpyxl.load_workbook(out)
            for name in (
                "KARSILASTIRMA-SRAVNENIE",
                "OZET-SVODKA",
                "FARKLAR-RAZNICA",
                "HESAP FARKI-SCHETA",
            ):
                self.assertIn(name, wb.sheetnames)
            # Dosya sayfalari sanitize edilmis etiketlerle
            self.assertTrue(any(s.startswith("A") or s == "A" for s in wb.sheetnames))
            self.assertEqual(result["match_count"], 1)
            self.assertEqual(result["mismatch_count"], 1)
            self.assertEqual(result["only1_count"], 1)
            self.assertEqual(result["only2_count"], 1)

    def test_combined_sheet_preserves_beleg_and_amount_chars(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "1.csv")
            f2 = os.path.join(tmp, "2.csv")
            text = "Aral Tankstelle Leonart Bewirtungskosten (nur MA)"
            _write_datev(f1, [_row("1,99", "S", "6640", "70064", "0103", "8796", text)])
            _write_datev(f2, [_row("1,99", "S", "6640", "70064", "0103", "8796", text)])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="D1", f2_label="D2")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            # Header + 2 data rows (pair)
            values = []
            for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
                values.append(row)
            self.assertEqual(len(values), 2)
            for row in values:
                # COMBINED: Kaynak, Satir, Umsatz, SH, Konto, Gegen, Datum, Belegfeld1, Text, Not
                self.assertEqual(row[2], "1,99")
                self.assertEqual(str(row[4]), "6640")
                self.assertEqual(str(row[5]), "70064")
                self.assertEqual(str(row[6]), "0103")
                self.assertEqual(str(row[7]), "8796")
                self.assertEqual(row[8], text)


class TestE2ECrossFormat(unittest.TestCase):
    def test_csv_vs_datev_xlsx_exact_fields(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "cross.xlsx")
            result = motor.build_report(
                SAMPLE_A, SAMPLE_DATEV_XLSX, out,
                f1_label="CSV", f2_label="XLSX",
            )
            self.assertGreaterEqual(result["match_count"], 1)
            e1, e2, *_ = motor.compare(SAMPLE_A, SAMPLE_DATEV_XLSX)
            m = [e for e in e1 if e["belegfeld1"] == "RE-1001"][0]
            self.assertTrue(m["matched"])
            self.assertEqual(m["pair_kind"], "MATCH")
            self.assertEqual(m["amt"], 1000.0)
            self.assertEqual(m["pair"]["amt"], 1000.0)
            self.assertEqual(m["text"], m["pair"]["text"])
            self.assertEqual(m["konto"], m["pair"]["konto"])

    def test_csv_vs_journal_xlsx_fields(self):
        e1, e2, *_ = motor.compare(SAMPLE_A, SAMPLE_JOURNAL)
        m = [e for e in e1 if e["belegfeld1"] == "RE-1001" and not e["netted"]][0]
        self.assertEqual(m["pair_kind"], "MATCH")
        self.assertEqual(m["amt"], m["pair"]["amt"])
        self.assertEqual(m["belegfeld1"], m["pair"]["belegfeld1"])
        self.assertEqual(m["datum"], m["pair"]["datum"])
        mm = [e for e in e1 if e["belegfeld1"] == "RE-2002" and not e["netted"]][0]
        self.assertEqual(mm["pair_kind"], "MISMATCH")
        self.assertNotEqual(mm["amt"], mm["pair"]["amt"])
        self.assertEqual(mm["belegfeld1"], mm["pair"]["belegfeld1"])


class TestE2EIdenticalRoundtrip(unittest.TestCase):
    def test_clone_file_zero_diffs(self):
        rows = [
            _row("1234,56", "S", "1350", "9000", "0101", "INV-001", "Barely Digital Gmbh"),
            _row("13,00", "S", "5900", "70002", "0201", "INV-SK-6675402", "Sammelkunde Dienstleister"),
            _row("0,01", "H", "1406", "6640", "0301", "CENT-1", "Rundung"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "x.csv")
            f2 = os.path.join(tmp, "y.csv")
            _write_datev(f1, rows)
            _write_datev(f2, rows)
            out = os.path.join(tmp, "clone.xlsx")
            result = motor.build_report(f1, f2, out, f1_label="X", f2_label="Y")
            self.assertEqual(result["match_count"], 3)
            self.assertEqual(result["mismatch_count"], 0)
            self.assertEqual(result["only1_count"], 0)
            self.assertEqual(result["only2_count"], 0)
            self.assertEqual(result["f1_red"], 0)
            self.assertEqual(result["f2_red"], 0)
            # Hesap farki yok
            self.assertEqual(result["diffs"], [])
            # Kaynak satirlardaki her alan rapor FARKLAR sayfasinda olmamali
            wb = openpyxl.load_workbook(out)
            ws = wb["FARKLAR-RAZNICA"]
            # Sadece header
            self.assertEqual(ws.max_row, 1)


class TestE2EDetectRejects(unittest.TestCase):
    def test_reject_unsupported_extension(self):
        with tempfile.NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8") as f:
            f.write("hello")
            path = f.name
        try:
            with self.assertRaises(motor.FormatError):
                motor.detect_format(path)
        finally:
            os.unlink(path)

    def test_reject_comparison_output_xlsx(self):
        # Onceki rapor formati kaynak degil
        path = os.path.join(EXCEL1, "KARSILASTIRMA_ISARETLI.xlsx")
        if not os.path.isfile(path):
            self.skipTest("KARSILASTIRMA_ISARETLI.xlsx yok")
        with self.assertRaises(motor.FormatError):
            motor.detect_format(path)


@unittest.skipUnless(os.path.isdir(MART_DIR), "mart klasoru yok")
class TestE2EMartRealFiles(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        files = os.listdir(MART_DIR)
        cls.csv_path = os.path.join(MART_DIR, [f for f in files if f.lower().endswith(".csv")][0])
        cls.xlsx_path = os.path.join(MART_DIR, [f for f in files if f.lower().endswith(".xlsx")][0])

    def test_formats_detected(self):
        self.assertEqual(motor.detect_format(self.csv_path), motor.FORMAT_DATEV_CSV)
        self.assertEqual(motor.detect_format(self.xlsx_path), motor.FORMAT_JOURNAL_XLSX)

    def test_load_counts_stable(self):
        e1 = motor.load(self.csv_path)
        e2 = motor.load(self.xlsx_path)
        self.assertEqual(len(e1), 689)
        self.assertEqual(len(e2), 812)

    def test_known_row_chars_intact_after_load(self):
        e1 = motor.load(self.csv_path)
        bew = [
            e for e in e1
            if e["belegfeld1"] == "8796"
            and abs(e["amt"] - 1.99) < 0.001
            and e["text"] == "Aral Tankstelle Leonart Bewirtungskosten (nur MA)"
        ]
        self.assertEqual(len(bew), 1)
        e = bew[0]
        self.assertEqual(e["umsatz"], "1,99")
        self.assertEqual(e["sh"], "S")
        self.assertEqual(e["konto"], "6640")
        self.assertEqual(e["gegenkonto"], "70064")
        self.assertEqual(e["datum"], "0103")
        self.assertEqual(e["belegfeld1"], "8796")
        self.assertEqual(e["amt"], 1.99)

    def test_known_row_matches_journal_exactly(self):
        e1, e2, *_ = motor.compare(self.csv_path, self.xlsx_path)
        hit = [
            e for e in e1
            if e["belegfeld1"] == "8796"
            and abs(e["amt"] - 1.99) < 0.001
            and e["text"] == "Aral Tankstelle Leonart Bewirtungskosten (nur MA)"
        ][0]
        self.assertTrue(hit["matched"])
        self.assertEqual(hit["pair_kind"], "MATCH")
        p = hit["pair"]
        self.assertEqual(p["amt"], 1.99)
        self.assertEqual(p["belegfeld1"], "8796")
        self.assertEqual(p["konto"], "6640")
        self.assertEqual(p["gegenkonto"], "70064")
        self.assertEqual(p["datum"], "0103")
        self.assertEqual(p["text"], hit["text"])

    def test_full_report_counts_and_excel(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "mart_e2e.xlsx")
            result = motor.build_report(
                self.csv_path, self.xlsx_path, out,
                f1_label="DATEV", f2_label="Journal",
            )
            self.assertGreaterEqual(result["match_count"], 600)
            self.assertEqual(result["f1_total"], 689)
            self.assertEqual(result["f2_total"], 812)
            # Netted kayitlar birlesik sayfada yok; her DATEV kaydi (net haric)
            # MATCH / MISMATCH / ONLY1 olur. net1 = storno cifti sayisi => 2*net1 kayit.
            self.assertEqual(
                result["match_count"] + result["mismatch_count"] + result["only1_count"],
                result["f1_total"] - 2 * result["net1"],
            )
            total_groups = (
                result["match_count"] + result["mismatch_count"]
                + result["only1_count"] + result["only2_count"]
            )
            self.assertGreater(total_groups, 600)

            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            # Bilinen satiri raporda bul (Umsatz 1,99 + Beleg 8796)
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[7] is not None and str(row[7]) == "8796" and row[2] in ("1,99", 1.99):
                    self.assertEqual(str(row[4]), "6640")
                    self.assertEqual(str(row[5]), "70064")
                    self.assertEqual(str(row[6]), "0103")
                    self.assertIn("Aral Tankstelle", str(row[8]))
                    found = True
                    break
            self.assertTrue(found, "8796 / 1,99 satiri raporda bulunamadi")

    def test_matched_pairs_amount_and_beleg_never_silently_altered(self):
        e1, e2, *_ = motor.compare(self.csv_path, self.xlsx_path)
        checked = 0
        for e in e1:
            if not e["matched"] or e["netted"] or e["pair_kind"] != "MATCH":
                continue
            p = e["pair"]
            self.assertEqual(e["amt"], p["amt"], msg=f"tutar degisti: {e['belegfeld1']}")
            # Beleg her iki tarafta da doluysa esit olmali (substring/recode haric)
            if e["belegfeld1"] and p["belegfeld1"]:
                # Tam eslesme katmaninda beleg normlari ayni olmali
                if e["tier"] and e["tier"].startswith("TAM ESLESME"):
                    self.assertEqual(e["beleg_n"], p["beleg_n"])
                    self.assertEqual(e["belegfeld1"], p["belegfeld1"])
            checked += 1
            if checked >= 200:
                break
        self.assertGreaterEqual(checked, 200)


@unittest.skipUnless(
    os.path.isfile(os.path.join(EXCEL1, "EXTF_Buchungsstapel_01012026_bis_30062026.csv"))
    and os.path.isfile(os.path.join(EXCEL1, "EXTF_MB_01-06.csv")),
    "excel1 DATEV cifti yok",
)
class TestE2ETwoDatevCsv(unittest.TestCase):
    def test_two_datev_csv_compare_runs(self):
        f1 = os.path.join(EXCEL1, "EXTF_Buchungsstapel_01012026_bis_30062026.csv")
        f2 = os.path.join(EXCEL1, "EXTF_MB_01-06.csv")
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "two_datev.xlsx")
            result = motor.build_report(f1, f2, out, f1_label="BS", f2_label="MB")
            self.assertGreater(result["f1_total"], 1000)
            self.assertGreater(result["f2_total"], 1000)
            self.assertTrue(os.path.isfile(out))
            # En az bir MATCH olmali
            self.assertGreater(result["match_count"], 100)
            wb = openpyxl.load_workbook(out)
            self.assertIn("KARSILASTIRMA-SRAVNENIE", wb.sheetnames)


class TestE2EUtf8AndSpecialChars(unittest.TestCase):
    def test_special_chars_in_text_survive_report(self):
        special = "Miete März / Talweg 4 - Dahlem & Co. KG №12"
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            # utf-8-sig yaz; motor latin-1 veya utf-8-sig denemeli
            # latin-1 ile yazamayiz (№); utf-8-sig kullan
            for path in (f1, f2):
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(["EXTF", "700", "21", "Buchungsstapel"])
                    w.writerow(DATEV_HEADER)
                    w.writerow(_row("985,00", "S", "6310", "1801", "0203", "MIETE", special))
            # detect/load utf-8-sig
            e = motor.load(f1)
            self.assertEqual(len(e), 1)
            self.assertEqual(e[0]["text"], special)
            out = os.path.join(tmp, "sp.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            texts = [row[8] for row in ws.iter_rows(min_row=2, values_only=True)]
            self.assertTrue(any(t == special for t in texts))


if __name__ == "__main__":
    unittest.main()
