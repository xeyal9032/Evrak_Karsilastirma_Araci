# -*- coding: utf-8 -*-
"""
Coklu E2E senaryolari: ters sira, fixture matrisi genis, stres, GUI thread-safe API.
"""
import os
import sys
import tempfile
import threading
import unittest
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(__file__))

import karsilastir_motor as motor
from fixture_paths import mart_dir, excel1_dir, has_mart
from helpers_datev import drow, write_datev_csv, write_datev_xlsx, write_journal_xlsx

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
MART = mart_dir() or ""
EXCEL1 = excel1_dir() or ""


class TestE2ESwapOrder(unittest.TestCase):
    def test_swap_files_preserves_symmetric_match_count(self):
        a = os.path.join(FIXTURES, "sample_a.csv")
        b = os.path.join(FIXTURES, "sample_b.csv")
        with tempfile.TemporaryDirectory() as tmp:
            r_ab = motor.build_report(a, b, os.path.join(tmp, "ab.xlsx"), f1_label="A", f2_label="B")
            r_ba = motor.build_report(b, a, os.path.join(tmp, "ba.xlsx"), f1_label="B", f2_label="A")
        # MATCH sayisi simetrik olmali
        self.assertEqual(r_ab["match_count"], r_ba["match_count"])
        self.assertEqual(r_ab["mismatch_count"], r_ba["mismatch_count"])
        # only1/only2 yer degistirir
        self.assertEqual(r_ab["only1_count"], r_ba["only2_count"])
        self.assertEqual(r_ab["only2_count"], r_ba["only1_count"])


class TestE2EMultiFormatMatrix(unittest.TestCase):
    def test_csv_xlsx_journal_triangle(self):
        """DATEV csv, DATEV xlsx, Journal ucgeni - hepsi hatasiz rapor uretir."""
        csv_p = os.path.join(FIXTURES, "sample_a.csv")
        xlsx_p = os.path.join(FIXTURES, "sample_datev.xlsx")
        journal_p = os.path.join(FIXTURES, "sample_journal.xlsx")
        combos = [
            (csv_p, xlsx_p),
            (csv_p, journal_p),
            (xlsx_p, journal_p),
            (xlsx_p, csv_p),
            (journal_p, csv_p),
            (journal_p, xlsx_p),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            for i, (f1, f2) in enumerate(combos):
                out = os.path.join(tmp, f"c{i}.xlsx")
                r = motor.build_report(f1, f2, out, f1_label="F1", f2_label="F2")
                self.assertTrue(os.path.isfile(out))
                self.assertGreater(r["f1_total"] + r["f2_total"], 0)
                wb = openpyxl.load_workbook(out)
                self.assertIn("KARSILASTIRMA-SRAVNENIE", wb.sheetnames)
                self.assertIn("OZET-SVODKA", wb.sheetnames)


class TestE2EGeneratedJournalVsDatev(unittest.TestCase):
    def test_generated_perfect_mirror(self):
        """Ayni ekonomik kayitlar DATEV CSV + Journal XLSX -> sifir kirmizi."""
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = os.path.join(tmp, "d.csv")
            j_path = os.path.join(tmp, "j.xlsx")
            write_datev_csv(csv_path, [
                drow("110,70", "S", "6530", "70064", "0103", "8796", "Kraftstoff"),
                drow("1,99", "S", "6640", "70064", "0103", "8796", "Bewirtung"),
                drow("1000,00", "H", "8400", "1200", "2001", "RE-H", "HabenSeite"),
            ])
            # Journal: H kaydi Soll bakisiyla (1200/8400)
            write_journal_xlsx(j_path, [
                (datetime(2026, 3, 1), "8796", "Kraftstoff", 110.70, 6530, 70064),
                (datetime(2026, 3, 1), "8796", "Bewirtung", 1.99, 6640, 70064),
                (datetime(2026, 1, 20), "RE-H", "HabenSeite", 1000.00, 1200, 8400),
            ])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(csv_path, j_path, out, f1_label="D", f2_label="J")
            self.assertEqual(r["match_count"], 3)
            self.assertEqual(r["mismatch_count"], 0)
            self.assertEqual(r["only1_count"], 0)
            self.assertEqual(r["only2_count"], 0)


class TestE2EStressManyRows(unittest.TestCase):
    def test_hundreds_of_unique_rows(self):
        rows1, rows2 = [], []
        for i in range(300):
            beleg = f"B{i:04d}"
            amt = f"{(i % 50) + 1},00"
            text = f"Buchung Text {i}"
            rows1.append(drow(amt, "S", "1000", "2000", "0101", beleg, text))
            # 280 ayni, 20 sadece f1; f2'de 20 ekstra
            if i < 280:
                rows2.append(drow(amt, "S", "1000", "2000", "0101", beleg, text))
        for i in range(280, 300):
            rows2.append(drow("9,00", "S", "3000", "4000", "0202", f"X{i}", f"Extra {i}"))
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, rows1)
            write_datev_csv(f2, rows2)
            out = os.path.join(tmp, "big.xlsx")
            r = motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            self.assertEqual(r["match_count"], 280)
            self.assertEqual(r["only1_count"], 20)
            self.assertEqual(r["only2_count"], 20)
            self.assertEqual(r["mismatch_count"], 0)


class TestE2EConcurrentBuild(unittest.TestCase):
    def test_two_threads_build_reports(self):
        a = os.path.join(FIXTURES, "sample_a.csv")
        b = os.path.join(FIXTURES, "sample_b.csv")
        results = [None, None]
        errors = []

        def worker(idx, out):
            try:
                results[idx] = motor.build_report(a, b, out, f1_label="A", f2_label="B")
            except Exception as ex:
                errors.append(ex)

        with tempfile.TemporaryDirectory() as tmp:
            t1 = threading.Thread(target=worker, args=(0, os.path.join(tmp, "t1.xlsx")))
            t2 = threading.Thread(target=worker, args=(1, os.path.join(tmp, "t2.xlsx")))
            t1.start(); t2.start()
            t1.join(); t2.join()
        self.assertEqual(errors, [])
        self.assertEqual(results[0]["match_count"], results[1]["match_count"])
        self.assertEqual(results[0]["only1_count"], results[1]["only1_count"])


class TestE2EAmountRoundtripAllMart(unittest.TestCase):
    @unittest.skipUnless(has_mart(), "mart yok")
    def test_every_loaded_umsatz_parses_to_amt(self):
        files = os.listdir(MART)
        csv_path = os.path.join(MART, [f for f in files if f.endswith(".csv")][0])
        xlsx_path = os.path.join(MART, [f for f in files if f.endswith(".xlsx")][0])
        for path in (csv_path, xlsx_path):
            for e in motor.load(path):
                self.assertEqual(motor.norm_amount(e["umsatz"]), e["amt"])
                self.assertIsInstance(e["amt"], float)
                # Negatif tutar (iade vb.) olabilir; sadece sayisal butunluk
                self.assertEqual(round(e["amt"], 2), e["amt"])


class TestE2EMatchFieldInvariantsMart(unittest.TestCase):
    @unittest.skipUnless(has_mart(), "mart yok")
    def test_invariants_on_all_pairs(self):
        files = os.listdir(MART)
        csv_path = os.path.join(MART, [f for f in files if f.endswith(".csv")][0])
        xlsx_path = os.path.join(MART, [f for f in files if f.endswith(".xlsx")][0])
        e1, e2, *_ = motor.compare(csv_path, xlsx_path)
        for e in e1 + e2:
            if e["pair"] is None:
                continue
            self.assertIs(e["pair"]["pair"], e)
            self.assertEqual(e["pair_kind"], e["pair"]["pair_kind"])
            if e["pair_kind"] == "MATCH":
                self.assertEqual(e["amt"], e["pair"]["amt"])
            if e["pair_kind"] == "MISMATCH":
                note = motor.diff_note(e, e["pair"])
                self.assertTrue(note)  # en az bir alan farki yazilmali


class TestE2EDatevXlsxRoundtrip(unittest.TestCase):
    def test_csv_to_xlsx_clone_all_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = os.path.join(tmp, "a.csv")
            xlsx_path = os.path.join(tmp, "a.xlsx")
            rows = [
                drow("21,48", "S", "5900", "70055", "0201", "INV-1", "Barely"),
                drow("13,00", "S", "5900", "70002", "0201", "INV-2", "Sammel"),
            ]
            write_datev_csv(csv_path, rows)
            write_datev_xlsx(xlsx_path, rows)
            r = motor.build_report(csv_path, xlsx_path, os.path.join(tmp, "r.xlsx"), f1_label="C", f2_label="X")
            self.assertEqual(r["match_count"], 2)
            self.assertEqual(r["only1_count"], 0)
            self.assertEqual(r["only2_count"], 0)


@unittest.skipUnless(
    os.path.isfile(os.path.join(EXCEL1, "EXTF_Buchungsstapel_01012026_bis_30062026.csv"))
    and os.path.isfile(os.path.join(EXCEL1, "EXTF_MB_01-06.csv")),
    "excel1 yok",
)
class TestE2ELargeDatevInvariants(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.f1 = os.path.join(EXCEL1, "EXTF_Buchungsstapel_01012026_bis_30062026.csv")
        cls.f2 = os.path.join(EXCEL1, "EXTF_MB_01-06.csv")
        cls.e1, cls.e2, cls.net1, cls.net2 = motor.compare(cls.f1, cls.f2)

    def test_pair_kind_symmetry(self):
        for e in self.e1:
            if e["pair"] is not None:
                self.assertEqual(e["pair_kind"], e["pair"]["pair_kind"])

    def test_no_match_with_unequal_amount(self):
        bad = [
            (e["belegfeld1"], e["amt"], e["pair"]["amt"])
            for e in self.e1
            if e.get("pair_kind") == "MATCH" and not e["netted"] and e["amt"] != e["pair"]["amt"]
        ]
        self.assertEqual(bad, [])

    def test_netting_entries_are_matched(self):
        for e in self.e1 + self.e2:
            if e["netted"]:
                self.assertTrue(e["matched"])
                self.assertTrue((e["tier"] or "").startswith("IC-NETLESME"))


class TestE2EGuiDesktopAndOpenHelpers(unittest.TestCase):
    def test_helpers(self):
        import karsilastir as gui
        self.assertTrue(os.path.isdir(gui.desktop_dir()))
        self.assertTrue(callable(gui.open_file))
        # FILETYPES csv/xlsx icermeli
        joined = " ".join(str(x) for x in gui.FILETYPES)
        self.assertIn("csv", joined.lower())
        self.assertIn("xlsx", joined.lower())


if __name__ == "__main__":
    unittest.main()
