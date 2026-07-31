# -*- coding: utf-8 -*-
"""HTML/PDF rapor, groups_to_rows, FILTRE sayfalari derin testleri."""
import json
import os
import re
import sys
import tempfile
import unittest
from unittest import mock

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import karsilastir_motor as motor
import report_extra
from tests.helpers_datev import drow, write_datev_csv


class TestGroupsToRows(unittest.TestCase):
    def test_groups_to_rows_match_mismatch_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("20,00", "S", "1", "2", "0202", "D", "diff"),
                drow("30,00", "S", "1", "2", "0303", "O1", "only1"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("99,00", "S", "1", "2", "0202", "D", "diff"),
                drow("40,00", "S", "1", "2", "0404", "O2", "only2"),
            ])
            e1, e2, *_ = motor.compare(f1, f2)
            groups = motor.build_combined_rows(e1, e2)
            rows = report_extra.groups_to_rows(groups, "A", "B", lang="en")
            statuses = [r["status"] for r in rows]
            self.assertEqual(statuses.count("MATCH"), 2)
            self.assertEqual(statuses.count("MISMATCH"), 2)
            self.assertEqual(statuses.count("ONLY1"), 2)
            self.assertEqual(statuses.count("ONLY2"), 2)
            mism = [r for r in rows if r["status"] == "MISMATCH"]
            self.assertEqual(mism[0]["note"], "")
            self.assertTrue(mism[1]["note"])
            only1 = [r for r in rows if r["status"] == "ONLY1"]
            self.assertEqual(only1[1]["note"], "MISSING")


class TestHtmlDeep(unittest.TestCase):
    def test_html_payload_json_statuses_and_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("20,00", "S", "1", "2", "0202", "D", "a"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("99,00", "S", "1", "2", "0202", "D", "a"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out, write_html=True)
            with open(r["html_path"], encoding="utf-8") as fh:
                text = fh.read()
            self.assertIn(f'id="c-match">{r["match_count"]}</b>', text)
            self.assertIn(f'id="c-mis">{r["mismatch_count"]}</b>', text)
            m = re.search(r"const ROWS = (\[.*?\]);", text, re.S)
            self.assertIsNotNone(m)
            payload = json.loads(m.group(1))
            self.assertTrue(any(x["status"] == "MATCH" for x in payload))
            self.assertTrue(any(x["status"] == "MISMATCH" for x in payload))

    def test_html_escapes_script_in_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            evil = '</script><script>alert(1)</script>'
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "E", evil)])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "E", evil)])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out, write_html=True)
            with open(r["html_path"], encoding="utf-8") as fh:
                text = fh.read()
            # Raw script-break must not appear; unicode escapes are fine
            self.assertNotIn("</script><script>", text)
            self.assertIn("\\u003c/script\\u003e", text)

    def test_html_diff_table_top10(self):
        diffs = [(f"K{i}", float(i), 0.0, float(i), 1, 0) for i in range(15)]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "t.html")
            report_extra.write_html_report(
                path, f1_path="a", f2_path="b", f1_label="A", f2_label="B",
                match_count=0, mismatch_count=0, only1_count=0, only2_count=0,
                groups=[], diffs=diffs, lang="en",
            )
            with open(path, encoding="utf-8") as fh:
                text = fh.read()
            self.assertEqual(text.count("<tr><td>K"), 10)

    def test_html_ui_follows_lang(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "tr.html")
            report_extra.write_html_report(
                path, f1_path="a", f2_path="b", f1_label="A", f2_label="B",
                match_count=1, mismatch_count=0, only1_count=0, only2_count=0,
                groups=[], lang="tr",
            )
            with open(path, encoding="utf-8") as fh:
                text = fh.read()
            self.assertIn('lang="tr"', text)
            self.assertIn("Eşleşme (sarı)", text)
            self.assertIn("Hesap farkları", text)
            self.assertIn(">Tümü<", text)
            self.assertNotIn("Match (yellow)", text)

            path_de = os.path.join(tmp, "de.html")
            report_extra.write_html_report(
                path_de, f1_path="a", f2_path="b", f1_label="A", f2_label="B",
                match_count=0, mismatch_count=0, only1_count=0, only2_count=0,
                groups=[], lang="de",
            )
            with open(path_de, encoding="utf-8") as fh:
                text_de = fh.read()
            self.assertIn("Treffer (gelb)", text_de)
            self.assertIn("Kontodifferenzen", text_de)


class TestPdfDeep(unittest.TestCase):
    def test_pdf_raises_without_reportlab(self):
        real_import = __import__

        def fake_import(name, *a, **k):
            if name == "reportlab" or name.startswith("reportlab."):
                raise ImportError("no reportlab")
            return real_import(name, *a, **k)

        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "x.pdf")
            with mock.patch("builtins.__import__", side_effect=fake_import):
                with self.assertRaises(RuntimeError) as cm:
                    report_extra.write_pdf_report(
                        path, f1_path="a", f2_path="b", f1_label="A", f2_label="B",
                        match_count=0, mismatch_count=0, only1_count=0, only2_count=0,
                        groups=[],
                    )
            self.assertIn("reportlab", str(cm.exception).lower())

    def test_pdf_unicode_turkish_renders(self):
        try:
            import reportlab  # noqa: F401
            from pypdf import PdfReader
        except ImportError:
            self.skipTest("reportlab/pypdf not installed")
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "tr.pdf")
            report_extra.write_pdf_report(
                path, f1_path="a.csv", f2_path="b.csv", f1_label="A", f2_label="B",
                match_count=3, mismatch_count=1, only1_count=0, only2_count=0,
                groups=[], lang="tr",
            )
            text = "".join((p.extract_text() or "") for p in PdfReader(path).pages)
            self.assertIn("Eşleşme", text)
            self.assertIn("Hesap farkları", text)
            self.assertNotIn("■■", text)


class TestFiltreAndFast(unittest.TestCase):
    def test_filtre_autofilter_absent_when_header_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            for name in ("FILTRE-KIRMIZI", "FILTRE-TURUNCU"):
                ws = wb[name]
                self.assertEqual(ws.max_row, 1)
                ref = ws.auto_filter.ref
                self.assertTrue(ref is None or ref == "" or ref.endswith("1"))

    def test_filtre_orange_note_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "D", "a")])
            write_datev_csv(f2, [drow("99,00", "S", "1", "2", "0101", "D", "a")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            ws = wb["FILTRE-TURUNCU"]
            self.assertGreaterEqual(ws.max_row, 3)
            note = ws.cell(3, 10).value  # partner row note (col J)
            self.assertTrue(note)
            self.assertIn("Umsatz", str(note))

    def test_fast_mode_keeps_filtre_and_combined(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B", detail_sheets=False)
            wb = openpyxl.load_workbook(out)
            names = set(wb.sheetnames)
            self.assertIn("KARSILASTIRMA-SRAVNENIE", names)
            self.assertIn("OZET-SVODKA", names)
            self.assertIn("HESAP FARKI-SCHETA", names)
            self.assertIn("FILTRE-KIRMIZI", names)
            self.assertIn("FILTRE-TURUNCU", names)
            self.assertNotIn("FARKLAR-RAZNICA", names)

    def test_combined_diff_note_only_on_partner_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "D", "a")])
            write_datev_csv(f2, [drow("99,00", "S", "1", "2", "0101", "D", "a")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out)
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            # Row 2 = e1, row 3 = e2 (1-indexed with header)
            note1 = ws.cell(2, 10).value
            note2 = ws.cell(3, 10).value
            self.assertTrue(note1 in (None, ""))
            self.assertTrue(note2)


if __name__ == "__main__":
    unittest.main()
