# -*- coding: utf-8 -*-
"""
Derin e2e: uretilen raporu satir satir dogrula, GUI smoke, mart mismatch notlari.
"""
import os
import sys
import tempfile
import unittest

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(__file__))

import karsilastir_motor as motor
from fixture_paths import mart_dir, excel1_dir, has_mart
from helpers_datev import drow, write_datev_csv

MART = mart_dir() or ""
FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")


def _rgb(cell):
    fill = cell.fill
    if not fill or not fill.fgColor or not fill.fgColor.rgb:
        return None
    r = str(fill.fgColor.rgb).upper()
    return r[-6:] if len(r) >= 6 else r


class TestE2EReportRowByRow(unittest.TestCase):
    def test_combined_sheet_mirrors_groups(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("20,00", "S", "1", "2", "0202", "D", "diff"),
                drow("30,00", "S", "3", "4", "0303", "O1", "only1"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("21,00", "S", "1", "2", "0202", "D", "diff"),
                drow("40,00", "S", "5", "6", "0404", "O2", "only2"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="L1", f2_label="L2")
            e1, e2, *_ = motor.compare(f1, f2)
            groups = motor.build_combined_rows(e1, e2)

            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            rows = list(ws.iter_rows(min_row=2))
            self.assertEqual(len(rows), 2 * len(groups))

            for gi, g in enumerate(groups):
                r1, r2 = rows[gi * 2], rows[gi * 2 + 1]
                kind = g["kind"]
                if kind == "MATCH":
                    self.assertEqual(_rgb(r1[0]), "FFFF00")
                    self.assertEqual(_rgb(r2[0]), "FFFF00")
                    self.assertEqual(str(r1[7].value), g["e1"]["belegfeld1"])
                    self.assertEqual(str(r2[7].value), g["e2"]["belegfeld1"])
                    self.assertEqual(motor.norm_amount(r1[2].value), g["e1"]["amt"])
                    self.assertEqual(motor.norm_amount(r2[2].value), g["e2"]["amt"])
                elif kind == "MISMATCH":
                    self.assertEqual(_rgb(r1[0]), "FFC000")
                    self.assertEqual(_rgb(r2[0]), "FFC000")
                    self.assertTrue(r2[9].value)  # diff note on second row
                elif kind == "ONLY1":
                    self.assertEqual(_rgb(r1[0]), "FF6B6B")
                    self.assertEqual(_rgb(r2[0]), "FF6B6B")
                    miss = str(r2[9].value).upper()
                    self.assertTrue(any(x in miss for x in ("EKSİK", "EKSIK", "MISSING", "FEHLT", "ОТСУТСТВУЕТ")))
                elif kind == "ONLY2":
                    self.assertEqual(_rgb(r1[0]), "FF6B6B")
                    self.assertEqual(_rgb(r2[0]), "FF6B6B")
                    miss = str(r1[9].value).upper()
                    self.assertTrue(any(x in miss for x in ("EKSİK", "EKSIK", "MISSING", "FEHLT", "ОТСУТСТВУЕТ")))

    def test_per_file_sheets_row_counts(self):
        a = os.path.join(FIXTURES, "sample_a.csv")
        b = os.path.join(FIXTURES, "sample_b.csv")
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(a, b, out, f1_label="FileA", f2_label="FileB")
            e1 = motor.load(a)
            e2 = motor.load(b)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb["FileA"].max_row - 1, len(e1))
            self.assertEqual(wb["FileB"].max_row - 1, len(e2))

    def test_headers_frozen(self):
        a = os.path.join(FIXTURES, "sample_a.csv")
        b = os.path.join(FIXTURES, "sample_b.csv")
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(a, b, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            for name in ("KARSILASTIRMA-SRAVNENIE", "FARKLAR-RAZNICA", "A", "B"):
                self.assertEqual(wb[name].freeze_panes, "A2")


@unittest.skipUnless(has_mart(), "mart yok")
class TestE2EMartMismatchNotesInReport(unittest.TestCase):
    def test_every_orange_pair_has_note(self):
        files = os.listdir(MART)
        cpath = os.path.join(MART, [f for f in files if f.endswith(".csv")][0])
        xpath = os.path.join(MART, [f for f in files if f.endswith(".xlsx")][0])
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "m.xlsx")
            r = motor.build_report(cpath, xpath, out, f1_label="D", f2_label="J")
            self.assertEqual(r["mismatch_count"], 6)
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            orange_notes = 0
            rows = list(ws.iter_rows(min_row=2))
            for i in range(0, len(rows), 2):
                if _rgb(rows[i][0]) == "FFC000":
                    self.assertEqual(_rgb(rows[i + 1][0]), "FFC000")
                    note = rows[i + 1][9].value
                    self.assertTrue(note and str(note).strip())
                    orange_notes += 1
            self.assertEqual(orange_notes, 6)


@unittest.skipUnless(has_mart(), "mart yok")
class TestE2EMartYellowAmountIntegrityInReport(unittest.TestCase):
    def test_all_yellow_pairs_amounts_equal_in_xlsx(self):
        files = os.listdir(MART)
        cpath = os.path.join(MART, [f for f in files if f.endswith(".csv")][0])
        xpath = os.path.join(MART, [f for f in files if f.endswith(".xlsx")][0])
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "m.xlsx")
            motor.build_report(cpath, xpath, out, f1_label="D", f2_label="J")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            rows = list(ws.iter_rows(min_row=2))
            bad = []
            for i in range(0, len(rows), 2):
                if _rgb(rows[i][0]) != "FFFF00" or _rgb(rows[i + 1][0]) != "FFFF00":
                    continue
                u1, u2 = rows[i][2].value, rows[i + 1][2].value
                if u1 in (None, "") or u2 in (None, ""):
                    continue  # ONLY tarafinda sari+kirmizi karisik degil burada
                a1, a2 = motor.norm_amount(u1), motor.norm_amount(u2)
                if a1 != a2:
                    bad.append((u1, u2, rows[i][7].value))
            self.assertEqual(bad, [])


class TestE2EGuiAppSmoke(unittest.TestCase):
    def test_app_constructs_without_mainloop(self):
        import tkinter as tk
        import karsilastir as gui

        # Headless-ish: create and destroy quickly
        try:
            root_check = tk.Tk()
            root_check.destroy()
        except tk.TclError:
            self.skipTest("Tk display unavailable")

        app = gui.App()
        try:
            self.assertEqual(app._busy, False)
            self.assertIsNone(app._fmt1)
            self.assertIsNone(app._fmt2)
            app.file1.set("x")
            app.file2.set("y")
            self.assertEqual(app.file1.get(), "x")
            # run_compare without valid files should warn paths - just ensure button exists
            self.assertTrue(str(app.compare_btn["state"]) in ("normal", "active"))
        finally:
            app.destroy()


class TestE2EDetectLabels(unittest.TestCase):
    def test_format_labels(self):
        self.assertEqual(motor.format_label(motor.FORMAT_DATEV_CSV), "DATEV CSV")
        self.assertEqual(motor.format_label(motor.FORMAT_DATEV_XLSX), "DATEV Excel")
        self.assertEqual(motor.format_label(motor.FORMAT_JOURNAL_XLSX), "Journal Excel")


class TestE2EHesapDiffConsistency(unittest.TestCase):
    def test_diffs_sorted_by_abs(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "111", "222", "0101", "A", "t"),
                drow("5000,00", "S", "333", "444", "0202", "B", "t2"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "111", "222", "0101", "A", "t"),
            ])
            e1, e2, *_ = motor.compare(f1, f2)
            diffs = motor.account_balance_diff(e1, e2)
            abs_vals = [abs(d[3]) for d in diffs]
            self.assertEqual(abs_vals, sorted(abs_vals, reverse=True))


class TestE2ECrossFieldBothDirections(unittest.TestCase):
    def test_text_to_beleg_and_beleg_to_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            # Direction 1 already covered; direction 2: f1 text has id, f2 beleg has id
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("500,00", "S", "4400", "1000", "1003", "", "RE9999-CROSS"),
            ])
            write_datev_csv(f2, [
                drow("400,00", "S", "4400", "1000", "1003", "RE9999-CROSS", "Other"),
            ])
            e1, e2, *_ = motor.compare(f1, f2)
            self.assertTrue(e1[0]["matched"])
            self.assertEqual(e1[0]["pair_kind"], "MISMATCH")
            self.assertEqual(e1[0]["amt"], 500.0)
            self.assertEqual(e1[0]["pair"]["amt"], 400.0)


if __name__ == "__main__":
    unittest.main()
