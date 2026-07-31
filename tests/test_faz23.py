# -*- coding: utf-8 -*-
"""Faz 2/3: HTML/PDF, archive, batch, fast mode."""
import os
import sys
import tempfile
import unittest
from io import StringIO

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import archive_db
import batch_compare
import karsilastir_motor as motor
from karsilastir import run_cli
from tests.helpers_datev import drow, write_datev_csv


class TestHtmlReport(unittest.TestCase):
    def test_html_contains_filters_and_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("20,00", "S", "1", "2", "0202", "D", "a"),
                drow("30,00", "S", "1", "2", "0303", "O1", "x"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "ok"),
                drow("99,00", "S", "1", "2", "0202", "D", "a"),
                drow("40,00", "S", "1", "2", "0404", "O2", "y"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out, write_html=True)
            self.assertTrue(r["html_path"])
            self.assertTrue(os.path.isfile(r["html_path"]))
            with open(r["html_path"], encoding="utf-8") as fh:
                text = fh.read()
            self.assertIn("data-filter", text)
            self.assertIn("ONLY1", text)
            self.assertIn("MISMATCH", text)
            self.assertIn(str(r["match_count"]), text)


class TestPdfReport(unittest.TestCase):
    def test_pdf_written_when_reportlab_available(self):
        try:
            import reportlab  # noqa: F401
        except ImportError:
            self.skipTest("reportlab not installed")
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            out = os.path.join(tmp, "r.xlsx")
            r = motor.build_report(f1, f2, out, write_pdf=True)
            self.assertTrue(r["pdf_path"])
            self.assertTrue(os.path.isfile(r["pdf_path"]))
            self.assertGreater(os.path.getsize(r["pdf_path"]), 500)


class TestFastMode(unittest.TestCase):
    def test_fast_skips_detail_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B", detail_sheets=False)
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertIn("KARSILASTIRMA-SRAVNENIE", wb.sheetnames)
            self.assertIn("HESAP FARKI-SCHETA", wb.sheetnames)
            self.assertNotIn("FARKLAR-RAZNICA", wb.sheetnames)
            self.assertNotIn("A", wb.sheetnames)
            self.assertNotIn("B", wb.sheetnames)


class TestArchiveDb(unittest.TestCase):
    def test_save_and_list(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = os.path.join(tmp, "a.db")
            result = {
                "match_count": 1, "mismatch_count": 0,
                "only1_count": 0, "only2_count": 0,
                "f1_total": 1, "f2_total": 1, "out_path": "x.xlsx",
            }
            rid = archive_db.save_comparison(
                result, file1="a.csv", file2="b.csv",
                f1_label="A", f2_label="B", elapsed_s=0.5,
                source="test", db_path=db,
            )
            self.assertIsInstance(rid, int)
            rows = archive_db.list_comparisons(db_path=db)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["match_count"], 1)
            self.assertEqual(rows[0]["file1"], "a.csv")


class TestBatch(unittest.TestCase):
    def test_pair_by_stem_and_run(self):
        with tempfile.TemporaryDirectory() as tmp:
            a = os.path.join(tmp, "A")
            b = os.path.join(tmp, "B")
            out = os.path.join(tmp, "out")
            os.makedirs(a)
            os.makedirs(b)
            write_datev_csv(os.path.join(a, "mart.csv"), [drow("10,00", "S", "1", "2", "0101", "X", "t")])
            write_datev_csv(os.path.join(b, "mart.csv"), [drow("10,00", "S", "1", "2", "0101", "X", "t")])
            write_datev_csv(os.path.join(a, "orphan.csv"), [drow("1,00", "S", "1", "2", "0101", "Y", "t")])
            pairs, ua, ub = batch_compare.pair_by_stem(a, b)
            self.assertEqual(len(pairs), 1)
            self.assertEqual(len(ua), 1)
            summary = batch_compare.run_batch(a, b, out, mode="stem", write_html=True)
            self.assertEqual(summary["pair_count"], 1)
            self.assertTrue(os.path.isdir(out))
            self.assertTrue(any(name.endswith(".xlsx") for name in os.listdir(out)))
            self.assertTrue(any(name.endswith(".html") for name in os.listdir(out)))


class TestCliExtras(unittest.TestCase):
    def test_cli_html_and_archive(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            out = os.path.join(tmp, "r.xlsx")
            db = os.path.join(tmp, "x.db")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "ok")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([f1, f2, "-o", out, "--html", "--archive", db, "--quiet"])
                stdout = sys.stdout.getvalue()
            finally:
                sys.stdout, sys.stderr = old
            self.assertEqual(code, 0)
            self.assertTrue(os.path.isfile(out.replace(".xlsx", ".html")))
            self.assertIn("HTML", stdout)
            rows = archive_db.list_comparisons(db_path=db)
            self.assertEqual(len(rows), 1)

    def test_cli_batch(self):
        with tempfile.TemporaryDirectory() as tmp:
            a = os.path.join(tmp, "A")
            b = os.path.join(tmp, "B")
            out = os.path.join(tmp, "out")
            os.makedirs(a)
            os.makedirs(b)
            write_datev_csv(os.path.join(a, "p1.csv"), [drow("10,00", "S", "1", "2", "0101", "X", "t")])
            write_datev_csv(os.path.join(b, "p1.csv"), [drow("10,00", "S", "1", "2", "0101", "X", "t")])
            old = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([a, b, "--batch", "-o", out, "--quiet"])
                stdout = sys.stdout.getvalue()
            finally:
                sys.stdout, sys.stderr = old
            self.assertEqual(code, 0)
            self.assertIn("pairs=1", stdout)


if __name__ == "__main__":
    unittest.main()
