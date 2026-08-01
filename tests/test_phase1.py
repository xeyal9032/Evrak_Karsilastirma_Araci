# -*- coding: utf-8 -*-
"""Faz 1: progress, filtre sayfalari, bos sheet, CLI."""
import os
import sys
import tempfile
import unittest
from datetime import datetime
from io import StringIO

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import karsilastir_motor as motor
from karsilastir import run_cli
from tests.helpers_datev import drow, write_datev_csv, write_datev_xlsx


class TestProgressCallback(unittest.TestCase):
    def test_progress_monotonic_and_reaches_100(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("10,00", "S", "100", "200", "0101", "B1", "t1")])
            write_datev_csv(f2, [drow("10,00", "S", "100", "200", "0101", "B1", "t1")])
            out = os.path.join(tmp, "out.xlsx")
            events = []

            def cb(pct, stage):
                events.append((pct, stage))

            motor.build_report(f1, f2, out, progress_cb=cb)
            self.assertTrue(events)
            pcts = [p for p, _ in events]
            self.assertEqual(pcts, sorted(pcts))
            self.assertEqual(pcts[-1], 100)
            self.assertGreaterEqual(len(events), 5)


class TestExcelFilters(unittest.TestCase):
    def test_durum_column_autofilter_and_filter_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "100", "200", "0101", "SAME", "ok"),
                drow("20,00", "S", "100", "200", "0101", "DIFF", "a"),
                drow("30,00", "S", "100", "200", "0101", "ONLYA", "x"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "100", "200", "0101", "SAME", "ok"),
                drow("99,00", "S", "100", "200", "0101", "DIFF", "a"),
                drow("40,00", "S", "100", "200", "0101", "ONLYB", "y"),
            ])
            out = os.path.join(tmp, "out.xlsx")
            r = motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            self.assertIn("FILTRE-KIRMIZI", wb.sheetnames)
            self.assertIn("FILTRE-TURUNCU", wb.sheetnames)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            headers = [c.value for c in ws[1]]
            self.assertEqual(headers[-1], "Durum")
            self.assertIsNotNone(ws.auto_filter.ref)
            statuses = {row[-1].value for row in ws.iter_rows(min_row=2)}
            self.assertTrue({"MATCH", "MISMATCH", "ONLY1", "ONLY2"} & statuses)

            red = wb["FILTRE-KIRMIZI"]
            red_status = {row[-1].value for row in red.iter_rows(min_row=2) if row[-1].value}
            self.assertTrue(red_status <= {"ONLY1", "ONLY2"})
            self.assertTrue(red_status)

            orange = wb["FILTRE-TURUNCU"]
            orange_status = {row[-1].value for row in orange.iter_rows(min_row=2) if row[-1].value}
            self.assertEqual(orange_status, {"MISMATCH"})
            self.assertGreaterEqual(r["mismatch_count"], 1)
            self.assertGreaterEqual(r["only1_count"] + r["only2_count"], 1)


class TestEmptyFirstSheet(unittest.TestCase):
    def test_journal_skips_empty_header_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "j.xlsx")
            wb = openpyxl.Workbook()
            ws0 = wb.active
            ws0.title = "Empty"
            ws0.append(["Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                        "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto"])
            # no data rows
            ws1 = wb.create_sheet("Data")
            ws1.append(["Journal"])
            ws1.append(["Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                        "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto"])
            ws1.append([datetime(2026, 1, 15), None, 1, "B1", "text", 12.5, "EUR", "1000", "2000", None, None])
            wb.save(path)
            wb.close()
            entries = motor.load_journal_xlsx(path)
            self.assertEqual(len(entries), 1)
            self.assertAlmostEqual(entries[0]["amt"], 12.5)

    def test_datev_xlsx_skips_empty_header_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "d.xlsx")
            wb = openpyxl.Workbook()
            ws0 = wb.active
            ws0.title = "Empty"
            ws0.append(["EXTF"])
            ws0.append([
                "Umsatz (ohne Soll/Haben-Kz)", "Soll/Haben-Kennzeichen", "WKZ Umsatz", "Kurs",
                "Basis-Umsatz", "WKZ Basis-Umsatz", "Konto", "Gegenkonto (ohne BU-Schluessel)",
                "BU-Schlüssel", "Belegdatum", "Belegfeld 1", "Belegfeld 2", "Skonto", "Buchungstext",
            ])
            ws1 = wb.create_sheet("Data")
            write_tmp = os.path.join(tmp, "full.xlsx")
            write_datev_xlsx(write_tmp, [drow("10,00", "S", "100", "200", "0101", "B1", "t")])
            # Build manually with empty first sheet
            src = openpyxl.load_workbook(write_tmp)
            for row in src.active.iter_rows(values_only=True):
                ws1.append(list(row))
            src.close()
            wb.save(path)
            wb.close()
            entries = motor.load_datev_xlsx(path)
            self.assertEqual(len(entries), 1)


class TestCLI(unittest.TestCase):
    def test_cli_success(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            out = os.path.join(tmp, "report.xlsx")
            write_datev_csv(f1, [drow("10,00", "S", "100", "200", "0101", "B1", "t")])
            write_datev_csv(f2, [drow("10,00", "S", "100", "200", "0101", "B1", "t")])
            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([f1, f2, "-o", out, "--lang", "en", "--quiet"])
            finally:
                stdout = sys.stdout.getvalue()
                sys.stdout, sys.stderr = old_out, old_err
            self.assertEqual(code, 0)
            self.assertTrue(os.path.isfile(out))
            self.assertIn("OK", stdout)
            self.assertIn("match=", stdout)

    def test_cli_missing_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            write_datev_csv(f1, [drow("10,00", "S", "100", "200", "0101", "B1", "t")])
            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = StringIO(), StringIO()
            try:
                code = run_cli([f1, os.path.join(tmp, "missing.csv"), "--quiet"])
            finally:
                sys.stdout, sys.stderr = old_out, old_err
            self.assertEqual(code, 1)


if __name__ == "__main__":
    unittest.main()
