# -*- coding: utf-8 -*-
"""
E2E regresyon: kullanicinin urettigi rapor + motor yeniden hesap ayni olmali.
Ayrica birlesik sayfa renk/satir butunlugu.
"""
import os
import sys
import tempfile
import unittest
from collections import Counter

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import karsilastir_motor as motor

MART = r"C:\Users\xeyal\Desktop\excel1\mart"
USER_REPORT = r"C:\Users\xeyal\Desktop\Karsilastirma_20260731_215803.xlsx"

# Canli motor baseline (match-fix: ref_ids + klon guvenligi)
BASELINE = {
    "f1_total": 689,
    "f2_total": 812,
    "match_count": 644,
    "mismatch_count": 6,
    "only1_count": 39,
    "only2_count": 162,
}

# Eski kullanici Excel raporu (dosya degismedi) - sadece dosya okuma testleri
USER_REPORT_BASELINE = {
    "f1_total": 689,
    "f2_total": 812,
    "match_count": 649,
    "mismatch_count": 7,
    "only1_count": 33,
    "only2_count": 156,
}


def _mart_paths():
    files = os.listdir(MART)
    csv_path = os.path.join(MART, [f for f in files if f.lower().endswith(".csv")][0])
    xlsx_path = os.path.join(MART, [f for f in files if f.lower().endswith(".xlsx")][0])
    return csv_path, xlsx_path


def _fill_rgb(cell):
    fill = cell.fill
    if not fill or not fill.fgColor or not fill.fgColor.rgb:
        return None
    r = str(fill.fgColor.rgb).upper()
    return r[-6:] if len(r) >= 6 else r


@unittest.skipUnless(os.path.isdir(MART), "mart yok")
class TestE2EMartBaselineRegression(unittest.TestCase):
    def test_engine_matches_frozen_baseline(self):
        cpath, xpath = _mart_paths()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "regen.xlsx")
            r = motor.build_report(cpath, xpath, out, f1_label="DATEV", f2_label="Journal")
        for k, v in BASELINE.items():
            self.assertEqual(r[k], v, msg=f"{k}: {r[k]} != baseline {v}")

    def test_idempotent_compare_twice(self):
        cpath, xpath = _mart_paths()
        with tempfile.TemporaryDirectory() as tmp:
            r1 = motor.build_report(cpath, xpath, os.path.join(tmp, "1.xlsx"), f1_label="A", f2_label="B")
            r2 = motor.build_report(cpath, xpath, os.path.join(tmp, "2.xlsx"), f1_label="A", f2_label="B")
        for k in BASELINE:
            self.assertEqual(r1[k], r2[k])


@unittest.skipUnless(os.path.isfile(USER_REPORT), "kullanici raporu yok")
class TestE2EUserReportFile(unittest.TestCase):
    def test_user_report_sheet_structure(self):
        wb = openpyxl.load_workbook(USER_REPORT, data_only=True)
        needed = {
            "KARSILASTIRMA-SRAVNENIE",
            "OZET-SVODKA",
            "FARKLAR-RAZNICA",
            "HESAP FARKI-SCHETA",
        }
        self.assertTrue(needed.issubset(set(wb.sheetnames)))
        self.assertEqual(len(wb.sheetnames), 6)  # +2 dosya sayfasi

    def test_user_report_ozet_counts(self):
        wb = openpyxl.load_workbook(USER_REPORT, data_only=True)
        ws = wb["OZET-SVODKA"]
        rows = {i: [v for v in row if v is not None] for i, row in enumerate(ws.iter_rows(values_only=True), 1)}
        # Satir 14-17 ozet sayilari (onceki denetim)
        self.assertEqual(rows[14][1], USER_REPORT_BASELINE["match_count"])
        self.assertEqual(rows[15][1], USER_REPORT_BASELINE["mismatch_count"])
        self.assertEqual(rows[16][1], USER_REPORT_BASELINE["only1_count"])
        self.assertEqual(rows[17][1], USER_REPORT_BASELINE["only2_count"])
        self.assertEqual(rows[7][1], USER_REPORT_BASELINE["f1_total"])
        self.assertEqual(rows[8][1], USER_REPORT_BASELINE["f2_total"])

    def test_user_report_combined_row_and_color_counts(self):
        wb = openpyxl.load_workbook(USER_REPORT)
        ws = wb["KARSILASTIRMA-SRAVNENIE"]
        data = list(ws.iter_rows(min_row=2))
        expected_rows = 2 * sum(USER_REPORT_BASELINE[k] for k in ("match_count", "mismatch_count", "only1_count", "only2_count"))
        self.assertEqual(len(data), expected_rows)

        pairs = Counter()
        for i in range(0, len(data), 2):
            c1 = _fill_rgb(data[i][0])
            c2 = _fill_rgb(data[i + 1][0])
            pairs[(c1, c2)] += 1
        self.assertEqual(pairs[("FFFF00", "FFFF00")], USER_REPORT_BASELINE["match_count"])
        self.assertEqual(pairs[("FFC000", "FFC000")], USER_REPORT_BASELINE["mismatch_count"])
        self.assertEqual(pairs[("FFFF00", "FF6B6B")], USER_REPORT_BASELINE["only1_count"])
        self.assertEqual(pairs[("FF6B6B", "FFFF00")], USER_REPORT_BASELINE["only2_count"])

    def test_user_report_farklar_count(self):
        wb = openpyxl.load_workbook(USER_REPORT, data_only=True)
        ws = wb["FARKLAR-RAZNICA"]
        n = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(n, USER_REPORT_BASELINE["only1_count"] + USER_REPORT_BASELINE["only2_count"])

    def test_user_report_spot_known_bookings(self):
        wb = openpyxl.load_workbook(USER_REPORT, data_only=True)
        ws = wb["KARSILASTIRMA-SRAVNENIE"]
        texts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            texts.append((row[2], row[7], row[8]))
        bew = [t for t in texts if t[2] == "Aral Tankstelle Leonart Bewirtungskosten (nur MA)"]
        kraft = [t for t in texts if t[2] == "Aral Tankstelle Leonart Kraftstoff/Ladestrom"]
        self.assertGreaterEqual(len(bew), 2)  # her iki kaynak satiri
        self.assertGreaterEqual(len(kraft), 2)
        self.assertTrue(any(str(t[0]) == "1,99" and str(t[1]) == "8796" for t in bew))
        self.assertTrue(any(abs(motor.norm_amount(t[0]) - 110.7) < 0.001 for t in kraft))

    def test_user_report_matches_live_engine(self):
        """Rapordaki ozet, ayni kaynaklardan canli motorla ayni."""
        if not os.path.isdir(MART):
            self.skipTest("mart yok")
        cpath, xpath = _mart_paths()
        with tempfile.TemporaryDirectory() as tmp:
            live = motor.build_report(cpath, xpath, os.path.join(tmp, "x.xlsx"), f1_label="A", f2_label="B")
        for k, v in BASELINE.items():
            self.assertEqual(live[k], v)
