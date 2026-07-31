# -*- coding: utf-8 -*-
"""Derin Excel rapor dogrulama: renkler, sayfalar, FARKLAR, ozet."""
import os
import sys
import tempfile
import unittest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
import karsilastir_motor as motor
from helpers_datev import drow, write_datev_csv


def _fill_rgb(cell):
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    c = fill.fgColor
    if c.type == "rgb" and c.rgb:
        return str(c.rgb).upper().lstrip("FF") if len(str(c.rgb)) == 8 else str(c.rgb).upper()
    if c.type == "theme":
        return None
    return str(c.rgb).upper() if c.rgb else None


class TestReportColors(unittest.TestCase):
    def test_combined_sheet_color_codes(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("20,00", "S", "1", "2", "0201", "D", "diff"),
                drow("30,00", "S", "3", "4", "0301", "O1", "only1"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("21,00", "S", "1", "2", "0201", "D", "diff"),
                drow("40,00", "S", "5", "6", "0401", "O2", "only2"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="L1", f2_label="L2")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]

            yellow, orange, red = set(), set(), set()
            for row in ws.iter_rows(min_row=2):
                rgb = _fill_rgb(row[0])
                if not rgb:
                    continue
                # openpyxl may return FFFF00 or 00FFFF00
                core = rgb[-6:]
                beleg = row[7].value
                if core == "FFFF00":
                    yellow.add(str(beleg) if beleg else "EMPTY")
                elif core == "FFC000":
                    orange.add(str(beleg) if beleg else "EMPTY")
                elif core == "FF6B6B":
                    red.add(str(beleg) if beleg else "EMPTY")

            self.assertIn("M", yellow)
            self.assertIn("D", orange)
            # ONLY tarafinda bos beleg + EKSIK notu kirmizi
            self.assertTrue(red)


class TestReportFarklarSheet(unittest.TestCase):
    def test_farklar_only_unmatched(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
                drow("30,00", "S", "3", "4", "0301", "ONLY1", "solo"),
            ])
            write_datev_csv(f2, [
                drow("10,00", "S", "1", "2", "0101", "M", "match"),
            ])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            ws = wb["FARKLAR-RAZNICA"]
            belegs = [row[7] for row in ws.iter_rows(min_row=2, values_only=True)]
            self.assertEqual(belegs, ["ONLY1"])
            self.assertNotIn("M", belegs)


class TestReportOzet(unittest.TestCase):
    def test_ozet_contains_paths_and_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "aaa.csv")
            f2 = os.path.join(tmp, "bbb.csv")
            write_datev_csv(f1, [drow("10,00", "S", "1", "2", "0101", "M", "t")])
            write_datev_csv(f2, [drow("10,00", "S", "1", "2", "0101", "M", "t")])
            out = os.path.join(tmp, "r.xlsx")
            result = motor.build_report(f1, f2, out, f1_label="AAA", f2_label="BBB")
            wb = openpyxl.load_workbook(out)
            ws = wb["OZET-SVODKA"]
            texts = []
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is not None:
                        texts.append(str(v))
            joined = " | ".join(texts)
            self.assertIn(f1, joined)
            self.assertIn(f2, joined)
            self.assertIn("SARI", joined.upper())
            self.assertEqual(result["match_count"], 1)


class TestReportSheetNames(unittest.TestCase):
    def test_long_and_invalid_labels_sanitized(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "A", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "A", "t")])
            out = os.path.join(tmp, "r.xlsx")
            bad1 = "A" * 50 + ":*?/\\"
            bad2 = "A" * 50 + ":*?/\\"
            motor.build_report(f1, f2, out, f1_label=bad1, f2_label=bad2)
            wb = openpyxl.load_workbook(out)
            for name in wb.sheetnames:
                self.assertLessEqual(len(name), 31)
                for ch in "[]:*?/\\":
                    self.assertNotIn(ch, name)
            # Iki dosya sayfasi benzersiz
            data_sheets = [s for s in wb.sheetnames if s not in (
                "KARSILASTIRMA-SRAVNENIE", "OZET-SVODKA", "FARKLAR-RAZNICA", "HESAP FARKI-SCHETA"
            )]
            self.assertEqual(len(data_sheets), 2)
            self.assertEqual(len(set(s.lower() for s in data_sheets)), 2)


class TestReportHesapFarkiHighlight(unittest.TestCase):
    def test_large_diff_highlighted_red(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("5000,00", "S", "7777", "8888", "0101", "BIG", "big")])
            write_datev_csv(f2, [drow("1,00", "S", "1111", "2222", "0202", "SML", "small")])
            out = os.path.join(tmp, "r.xlsx")
            result = motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            self.assertTrue(any(abs(d[3]) >= 1000 for d in result["diffs"]))
            wb = openpyxl.load_workbook(out)
            ws = wb["HESAP FARKI-SCHETA"]
            red_rows = 0
            for row in ws.iter_rows(min_row=2):
                rgb = _fill_rgb(row[0])
                if rgb and rgb[-6:] == "FF6B6B":
                    red_rows += 1
            self.assertGreaterEqual(red_rows, 1)


class TestMismatchNoteInReport(unittest.TestCase):
    def test_orange_row_has_diff_note(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("100,00", "S", "1", "2", "0101", "X", "Same")])
            write_datev_csv(f2, [drow("200,00", "S", "1", "2", "0101", "X", "Same")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            notes = [row[9] for row in ws.iter_rows(min_row=2, values_only=True) if row[9]]
            self.assertTrue(any("100,00" in str(n) and "200,00" in str(n) for n in notes))


if __name__ == "__main__":
    unittest.main()
