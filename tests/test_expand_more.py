# -*- coding: utf-8 -*-
"""Genisletilmis testler: quoted CSV, coklu sheet, yesil netleme raporu, diff_note."""
import csv
import os
import sys
import tempfile
import unittest
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(__file__))

import karsilastir_motor as motor
from helpers_datev import DATEV_HEADER, drow, write_datev_csv, write_journal_xlsx


def _rgb(cell):
    fill = cell.fill
    if not fill or not fill.fgColor or not fill.fgColor.rgb:
        return None
    r = str(fill.fgColor.rgb).upper()
    return r[-6:] if len(r) >= 6 else r


class TestQuotedCsvFields(unittest.TestCase):
    def test_quoted_semicolon_inside_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "q.csv")
            with open(path, "w", encoding="latin-1", newline="") as f:
                w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
                w.writerow(["EXTF", "700", "21", "Buchungsstapel"])
                w.writerow(DATEV_HEADER)
                # text icinde noktali virgul benzeri icerik
                w.writerow([
                    "12,34", "S", "", "", "", "", "1000", "2000", "", "0101",
                    "BELEG;1", "", "", "Firma GmbH; Abteilung A",
                ])
            e = motor.load(path)
            self.assertEqual(len(e), 1)
            self.assertEqual(e[0]["belegfeld1"], "BELEG;1")
            self.assertEqual(e[0]["text"], "Firma GmbH; Abteilung A")
            self.assertEqual(e[0]["amt"], 12.34)


class TestJournalOnSecondSheet(unittest.TestCase):
    def test_detect_and_load_journal_not_on_first_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "multi.xlsx")
            wb = openpyxl.Workbook()
            ws0 = wb.active
            ws0.title = "Cover"
            ws0.append(["Not a journal"])
            ws1 = wb.create_sheet("Journal")
            ws1.append(["Titel"])
            ws1.append([
                "Belegdat.", "Zusatzang.", "Periode", "Belegnr.", "Buchungstext",
                "Betrag", "Whrg", "Sollkto", "Habenkto", "USt-EUR", "USt Kto",
            ])
            ws1.append([datetime(2026, 5, 5), None, 5, "J99", "SecondSheet", 42.5, "EUR", 10, 20, None, None])
            wb.save(path)
            wb.close()
            self.assertEqual(motor.detect_format(path), motor.FORMAT_JOURNAL_XLSX)
            e = motor.load(path)
            self.assertEqual(len(e), 1)
            self.assertEqual(e[0]["belegfeld1"], "J99")
            self.assertEqual(e[0]["amt"], 42.5)
            self.assertEqual(e[0]["datum"], "0505")


class TestDatevOnSecondSheet(unittest.TestCase):
    def test_datev_xlsx_second_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "d.xlsx")
            wb = openpyxl.Workbook()
            wb.active.title = "Info"
            wb.active.append(["hello"])
            ws = wb.create_sheet("Export")
            ws.append(["EXTF", "700"])
            ws.append([
                "Umsatz", "Soll/Haben-Kennzeichen", "WKZ Umsatz", "Kurs", "Basisumsatz",
                "WKZ Basisumsatz", "Konto", "Gegenkonto", "BU-Schluessel", "Belegdatum",
                "Belegfeld 1", "Belegfeld 2", "Skonto", "Buchungstext",
            ])
            ws.append(drow("9,99", "S", "1", "2", "1101", "S2", "Second"))
            wb.save(path)
            wb.close()
            self.assertEqual(motor.detect_format(path), motor.FORMAT_DATEV_XLSX)
            e = motor.load(path)
            self.assertEqual(e[0]["belegfeld1"], "S2")
            self.assertEqual(e[0]["amt"], 9.99)


class TestNettedGreenInReport(unittest.TestCase):
    def test_file_sheet_shows_green_for_netted(self):
        rows = [
            drow("100,00", "S", "2000", "2800", "0503", "RE-S", "Original Buchung"),
            drow("100,00", "S", "2800", "2000", "0503", "RE-S", "<Storno> Original Buchung"),
            drow("50,00", "S", "1000", "2000", "0603", "OK", "Normal"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, rows)
            write_datev_csv(f2, [drow("50,00", "S", "1000", "2000", "0603", "OK", "Normal")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="Src", f2_label="Dst")
            wb = openpyxl.load_workbook(out)
            ws = wb["Src"]
            greens = 0
            for row in ws.iter_rows(min_row=2):
                if _rgb(row[0]) == "A9D18E":
                    greens += 1
                    status = str(row[8].value)
                    self.assertIn("IC-NETLESME", status)
            self.assertEqual(greens, 2)


class TestDiffNoteEdges(unittest.TestCase):
    def test_identical_entries_empty_note(self):
        a = motor.make_entry(1, "1,00", "S", "1", "2", "0101", "X", "T", 1.0)
        b = motor.make_entry(2, "1,00", "S", "1", "2", "0101", "X", "T", 1.0)
        self.assertEqual(motor.diff_note(a, b), "")

    def test_none_partner_empty(self):
        a = motor.make_entry(1, "1,00", "S", "1", "2", "0101", "X", "T", 1.0)
        self.assertEqual(motor.diff_note(a, None), "")
        self.assertEqual(motor.diff_note(None, a), "")

    def test_multiple_field_changes(self):
        a = motor.make_entry(1, "1,00", "S", "1", "2", "0101", "X", "T1", 1.0)
        b = motor.make_entry(2, "2,00", "S", "9", "2", "0101", "Y", "T2", 2.0)
        note = motor.diff_note(a, b)
        self.assertIn("Umsatz", note)
        self.assertIn("Konto", note)
        self.assertIn("Belegfeld1", note)
        self.assertIn("Buchungstext", note)


class TestSingleRowFiles(unittest.TestCase):
    def test_one_vs_one_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("7,00", "S", "1", "2", "0101", "ONLY", "One")])
            write_datev_csv(f2, [drow("7,00", "S", "1", "2", "0101", "ONLY", "One")])
            r = motor.build_report(f1, f2, os.path.join(tmp, "r.xlsx"), f1_label="A", f2_label="B")
            self.assertEqual(r["match_count"], 1)
            self.assertEqual(r["f1_total"], 1)
            self.assertEqual(r["f2_total"], 1)

    def test_one_vs_one_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("7,00", "S", "1", "2", "0101", "A", "One")])
            write_datev_csv(f2, [drow("8,00", "S", "3", "4", "0202", "B", "Two")])
            r = motor.build_report(f1, f2, os.path.join(tmp, "r.xlsx"), f1_label="A", f2_label="B")
            self.assertEqual(r["only1_count"], 1)
            self.assertEqual(r["only2_count"], 1)
            self.assertEqual(r["match_count"], 0)


class TestUnicodeBelegAndText(unittest.TestCase):
    def test_unicode_survives_utf8_csv_and_report(self):
        text = "Ödeme / Zahlung — Café №7"
        beleg = "Fatura-Ü-001"
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("15,00", "S", "1", "2", "0101", beleg, text)], encoding="utf-8-sig")
            write_datev_csv(f2, [drow("15,00", "S", "1", "2", "0101", beleg, text)], encoding="utf-8-sig")
            e = motor.load(f1)
            self.assertEqual(e[0]["text"], text)
            self.assertEqual(e[0]["belegfeld1"], beleg)
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            ws = wb["KARSILASTIRMA-SRAVNENIE"]
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[8] == text and str(row[7]) == beleg:
                    found = True
            self.assertTrue(found)


class TestOzetLegendFills(unittest.TestCase):
    def test_legend_cells_colored(self):
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, [drow("1,00", "S", "1", "2", "0101", "A", "t")])
            write_datev_csv(f2, [drow("1,00", "S", "1", "2", "0101", "A", "t")])
            out = os.path.join(tmp, "r.xlsx")
            motor.build_report(f1, f2, out, f1_label="A", f2_label="B")
            wb = openpyxl.load_workbook(out)
            ws = wb["OZET-SVODKA"]
            colors = set()
            for row in ws.iter_rows(min_row=1, max_col=1):
                c = _rgb(row[0])
                if c:
                    colors.add(c)
            # Sari, turuncu, yesil, kirmizi legend
            self.assertTrue({"FFFF00", "FFC000", "A9D18E", "FF6B6B"}.issubset(colors))


class TestBuildCombinedSkipsNetted(unittest.TestCase):
    def test_netted_not_in_combined_groups(self):
        rows = [
            drow("100,00", "S", "2000", "2800", "0503", "RE-S", "Original"),
            drow("100,00", "S", "2800", "2000", "0503", "RE-S", "<Storno> Original"),
            drow("50,00", "S", "1000", "2000", "0603", "OK", "Normal"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            f1 = os.path.join(tmp, "a.csv")
            f2 = os.path.join(tmp, "b.csv")
            write_datev_csv(f1, rows)
            write_datev_csv(f2, [drow("50,00", "S", "1000", "2000", "0603", "OK", "Normal")])
            e1, e2, n1, n2 = motor.compare(f1, f2)
            groups = motor.build_combined_rows(e1, e2)
            self.assertEqual(n1, 1)
            self.assertEqual(len(groups), 1)
            self.assertEqual(groups[0]["kind"], "MATCH")
            self.assertEqual(groups[0]["e1"]["belegfeld1"], "OK")


class TestKeyValueEmptyGuards(unittest.TestCase):
    def test_empty_beleg_and_text_return_none(self):
        e = motor.make_entry(1, "1,00", "S", "1", "2", "0101", "", "", 1.0)
        self.assertIsNone(motor.key_value1(e))
        self.assertIsNone(motor.key_value2(e))


class TestSanitizeSheetEdge(unittest.TestCase):
    def test_empty_name_uses_fallback(self):
        taken = set()
        name = motor.sanitize_sheet_name("   ", "Fallback1", taken)
        self.assertEqual(name, "Fallback1")

    def test_only_invalid_chars(self):
        taken = set()
        name = motor.sanitize_sheet_name(":::***", "FB", taken)
        self.assertEqual(name, "FB")


class TestGuiDetectOnPickLogic(unittest.TestCase):
    def test_detect_labels_for_status(self):
        # GUI'nin gosterecegi etiketler
        a = os.path.join(os.path.dirname(__file__), "fixtures", "sample_a.csv")
        j = os.path.join(os.path.dirname(__file__), "fixtures", "sample_journal.xlsx")
        self.assertEqual(motor.format_label(motor.detect_format(a)), "DATEV CSV")
        self.assertEqual(motor.format_label(motor.detect_format(j)), "Journal Excel")


if __name__ == "__main__":
    unittest.main()
