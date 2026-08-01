# -*- coding: utf-8 -*-
"""
DATEV / Journal evrak karsilastirma motoru.
Iki muhasebe export dosyasini (DATEV CSV/Excel veya Journal Excel) otomatik
tanir, ortak kayit bicimine cevirir, satir satir karsilastirir ve renkli bir
Excel raporu uretir.

Desteklenen kaynak formatlar:
  - DATEV Buchungsstapel CSV (EXTF)
  - DATEV Buchungsstapel Excel (.xlsx)
  - Journal Excel (Belegdat. / Sollkto / Habenkto / Betrag)

Kullandigi mantik (elle yapilan analizden gelistirildi):
  1. Tier 1: Umsatz + Soll/Haben + Konto + Belegdatum + Belegfeld1(normalize) -> tam eslesme
  2. Tier 2: ayni ama Belegfeld1 yerine Buchungstext(ilk 60 karakter) kullanilir
  3. Tier 3: Konto'yu da yok sayar (hesap yeniden kodlamasini yakalar)
  4. Tier 4: sadece Umsatz+SH+Konto+Datum (son care, supheli eslesme olarak isaretlenir)
  0. Herhangi bir tier'dan once: dosya icinde kendi kendini iptal eden
     <Storno> + esi ciftleri tespit edilip netlenir (mukerrer kayit + iptali
     gercek fark sanmamak icin).
"""
import csv
import io
import os
import re
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

# ---- DATEV sutun basliklari (iki farkli export varyanti da desteklenir) ----
COL_NAMES = {
    "umsatz": ["Umsatz (ohne Soll/Haben-Kz)", "Umsatz"],
    "sh": ["Soll/Haben-Kennzeichen"],
    "konto": ["Konto"],
    "gegenkonto": ["Gegenkonto (ohne BU-Schluessel)", "Gegenkonto (ohne BU-Schl�ssel)", "Gegenkonto"],
    "datum": ["Belegdatum"],
    "belegfeld1": ["Belegfeld 1"],
    "buchungstext": ["Buchungstext"],
}


class FormatError(ValueError):
    """Dosya desteklenen muhasebe formatinda degil / Файл не в поддерживаемом формате."""
    pass


FORMAT_DATEV_CSV = "datev_csv"
FORMAT_DATEV_XLSX = "datev_xlsx"
FORMAT_JOURNAL_XLSX = "journal_xlsx"

FORMAT_LABELS = {
    FORMAT_DATEV_CSV: "DATEV CSV",
    FORMAT_DATEV_XLSX: "DATEV Excel",
    FORMAT_JOURNAL_XLSX: "Journal Excel",
}

JOURNAL_REQUIRED = ("Belegdat.", "Sollkto", "Habenkto", "Betrag", "Belegnr.", "Buchungstext")


def find_columns(header_row, fpath=""):
    idx = {}
    header_clean = [_cell_str(h) for h in header_row]
    for key, candidates in COL_NAMES.items():
        found = None
        # 1) tam eslesme (herhangi bir aday ile) / точное совпадение
        for cand in candidates:
            for i, h in enumerate(header_clean):
                if h == cand:
                    found = i
                    break
            if found is not None:
                break
        # 2) esnek arama: parantezden onceki kismi baslangic olarak eslestir
        #    (her sutun icin gecerli, sadece Gegenkonto icin degil)
        if found is None:
            base = candidates[0].split(" (")[0]
            for i, h in enumerate(header_clean):
                if h.startswith(base) and base:
                    found = i
                    break
        if found is None:
            raise FormatError(
                f"'{fpath}' dosyasinda beklenen sutun bulunamadi: {key} (aranan: {candidates}).\n"
                f"Bu dosya beklenen DATEV Buchungsstapel formatinda degil olabilir.\n\n"
                f"В файле '{fpath}' не найден ожидаемый столбец: {key} (искали: {candidates}).\n"
                f"Возможно, файл не в формате DATEV Buchungsstapel."
            )
        idx[key] = found
    return idx


def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip().strip('"')


def norm_amount(s):
    """Parse German DATEV or English decimal amounts to float (2 decimals).

    Examples:
      1.234.567,89 / 1,99 / 1000,00  (DE)
      12.34 / 1,234.56 / 1234.56     (EN)
      1.234                          (DE thousands, no cents)
    """
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return round(float(s), 2)
    s = _cell_str(s)
    if not s:
        return None
    s = s.replace(" ", "").replace("\u00a0", "")
    # Keep sign if present
    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:]
    elif s.startswith("+"):
        s = s[1:]
    if not s:
        return None
    try:
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                # 1.234,56
                s = s.replace(".", "").replace(",", ".")
            else:
                # 1,234.56
                s = s.replace(",", "")
        elif "," in s:
            # Only comma: DE decimal, or EN thousands (1,234)
            if re.fullmatch(r"\d{1,3}(,\d{3})+", s):
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")
        elif "." in s:
            # Only dot: DE thousands (1.234 / 1.234.567) or EN decimal (12.34)
            if re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
                s = s.replace(".", "")
            # else keep as EN decimal
        val = round(float(s), 2)
        return -val if neg else val
    except ValueError:
        return None


def format_umsatz(amt):
    return f"{amt:.2f}".replace(".", ",")


def format_belegdatum(v):
    if isinstance(v, datetime):
        return f"{v.day:02d}{v.month:02d}"
    if isinstance(v, date):
        return f"{v.day:02d}{v.month:02d}"
    s = _cell_str(v)
    if len(s) == 4 and s.isdigit():
        return s
    # Excel'den gelen '01.03.2026' / '1.3.2026'
    for sep in (".", "/", "-"):
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                return f"{int(parts[0]):02d}{int(parts[1]):02d}"
    return s


# Tamami bilimsel notasyon olan Belegfeld (Excel); "RE2026" gibi normal
# numaralardaki "E2026" yanlis pozitif olmasin diye ^...$ kullanilir.
_SCI_BELEG = re.compile(r"^[+-]?\d+([.,]\d+)?[Ee][+-]?\d+$")


def norm_belegfeld(s):
    """Belegfeld1 normalize. Excel bilimsel notasyonu (1,22001E+11 / 122001E+11)
    guvenilir bir belge no DEGILDIR - bos dondurur (yanlis eslesmeyi onler)."""
    s = _cell_str(s).strip()
    raw = s.replace(" ", "")
    if raw and _SCI_BELEG.match(raw.replace(",", ".")):
        return ""
    # "122001E+11" (noktasiz Excel kalintisi) - digit+E+digit
    if raw and re.match(r"^\d+[Ee][+-]?\d+$", raw):
        return ""
    return raw.replace(",", "").replace(".", "").upper()


def extract_ref_ids(*parts, min_len=7):
    """Metin/beleg icindeki uzun rakam dizilerini (fatura/musteri no) cikarir.
    Hem bosluklu orijinal hem de bosluksuz varyant taranir:
    - '0000119564441 0020744324955' -> iki ayri no
    - '2026 130000011' ve '2026130000011' -> ortak no yakalanir
    Basdaki sifirlar atilir: 0000119564441 -> 119564441."""
    ids = set()
    for p in parts:
        s = _cell_str(p)
        variants = (s, re.sub(r"[\s\-_/]+", "", s))
        for v in variants:
            for m in re.findall(r"\d{" + str(min_len) + r",}", v):
                norm = m.lstrip("0") or m
                if len(norm) >= min_len:
                    ids.add(norm)
    return ids


def norm_text(s):
    s = _cell_str(s)
    s = s.replace("<Storno> ", "").replace("<Storno>", "").strip()
    return s[:60].upper()


def make_entry(line, umsatz_raw, sh, konto, gegenkonto, datum, belegfeld1, text, amt):
    belegfeld1_raw = _cell_str(belegfeld1)
    text_raw = _cell_str(text)
    return {
        "line": line,
        "umsatz": _cell_str(umsatz_raw) if umsatz_raw is not None and not isinstance(umsatz_raw, (int, float)) else format_umsatz(amt),
        "sh": _cell_str(sh),
        "konto": _cell_str(konto),
        "gegenkonto": _cell_str(gegenkonto),
        "datum": _cell_str(datum) if not isinstance(datum, (date, datetime)) else format_belegdatum(datum),
        "belegfeld1": belegfeld1_raw,
        "text": text_raw,
        "amt": amt,
        "beleg_n": norm_belegfeld(belegfeld1_raw),
        "text_n": norm_text(text_raw),
        "ref_ids": extract_ref_ids(belegfeld1_raw, text_raw),
        "matched": False,
        "tier": None,
        "netted": False,
        "pair": None,
        "pair_kind": None,
    }


def _header_has_datev(header_cells):
    names = {_cell_str(h) for h in header_cells if _cell_str(h)}
    has_umsatz = any(n == "Umsatz" or n.startswith("Umsatz") for n in names)
    has_sh = "Soll/Haben-Kennzeichen" in names
    has_konto = "Konto" in names
    return has_umsatz and has_sh and has_konto


def _header_has_journal(header_cells):
    names = {_cell_str(h) for h in header_cells if _cell_str(h)}
    return all(req in names for req in ("Belegdat.", "Sollkto", "Habenkto", "Betrag"))


def _read_csv_rows(fpath):
    try:
        with open(fpath, "rb") as f:
            raw = f.read()
    except OSError as ex:
        raise FormatError(
            f"'{fpath}' dosyasi okunamadi.\n"
            f"Не удалось прочитать файл '{fpath}'.\n\n{ex}"
        )
    if not raw:
        raise FormatError(
            f"'{fpath}' dosyasi bos.\n"
            f"Файл '{fpath}' пустой."
        )
    # BOM varsa utf-8; aksi halde once utf-8 dene (latin-1 her bayti
    # "basarili" saydigi icin once deneyemeyiz - aksi halde UTF-8 harfler bozulur).
    if raw.startswith(b"\xef\xbb\xbf"):
        encodings = ("utf-8-sig",)
    else:
        encodings = ("utf-8", "cp1252", "latin-1")
    last_err = None
    text = None
    for enc in encodings:
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError as ex:
            last_err = ex
            continue
    if text is None:
        raise FormatError(
            f"'{fpath}' dosyasi okunamadi (metin kodlamasi sorunu).\n"
            f"Не удалось прочитать файл '{fpath}' (проблема кодировки).\n\n{last_err}"
        )
    try:
        return list(csv.reader(io.StringIO(text), delimiter=";"))
    except csv.Error as ex:
        raise FormatError(
            f"'{fpath}' dosyasi CSV olarak ayristirilamadi.\n"
            f"Не удалось разобрать файл '{fpath}' как CSV.\n\n{ex}"
        )


def _xlsx_sheet_rows(fpath):
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    except Exception as ex:
        raise FormatError(
            f"'{fpath}' Excel dosyasi acilamadi.\n"
            f"Не удалось открыть Excel-файл '{fpath}'.\n\n{ex}"
        )
    try:
        sheets = []
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            sheets.append((ws.title, rows))
        return sheets
    finally:
        wb.close()


def detect_format(fpath):
    """Dosya yolundan desteklenen kaynak formatini dondurur."""
    if not os.path.isfile(fpath):
        raise FileNotFoundError(fpath)
    ext = os.path.splitext(fpath)[1].lower()
    if ext == ".csv":
        rows = _read_csv_rows(fpath)
        if len(rows) >= 2 and _header_has_datev(rows[1]):
            return FORMAT_DATEV_CSV
        if len(rows) >= 1 and _header_has_datev(rows[0]):
            return FORMAT_DATEV_CSV
        raise FormatError(
            f"'{fpath}' CSV dosyasi DATEV Buchungsstapel olarak taninamadi "
            f"(beklenen sutunlar: Umsatz, Soll/Haben-Kennzeichen, Konto, ...).\n\n"
            f"Файл '{fpath}' не распознан как DATEV Buchungsstapel."
        )
    if ext in (".xlsx", ".xlsm"):
        for _title, rows in _xlsx_sheet_rows(fpath):
            for row in rows[:30]:
                if not row:
                    continue
                if _header_has_journal(row):
                    return FORMAT_JOURNAL_XLSX
                if _header_has_datev(row):
                    return FORMAT_DATEV_XLSX
        raise FormatError(
            f"'{fpath}' Excel dosyasi desteklenen bir kaynak format olarak taninamadi.\n"
            f"Desteklenen: DATEV Buchungsstapel veya Journal (Belegdat./Sollkto/Habenkto).\n\n"
            f"Файл Excel '{fpath}' не распознан. Поддерживаются: DATEV Buchungsstapel или Journal."
        )
    raise FormatError(
        f"'{fpath}' desteklenmeyen uzanti ({ext or 'yok'}). "
        f"Kullanin: .csv veya .xlsx\n\n"
        f"Неподдерживаемое расширение '{fpath}' ({ext or 'нет'}). Используйте .csv или .xlsx."
    )


def format_label(fmt):
    return FORMAT_LABELS.get(fmt, fmt)


def _entries_from_datev_table(rows, fpath, header_idx):
    if header_idx is None or header_idx >= len(rows):
        raise FormatError(
            f"'{fpath}' dosyasinda DATEV baslik satiri bulunamadi.\n"
            f"В файле '{fpath}' не найдена строка заголовка DATEV."
        )
    header_row = rows[header_idx]
    col = find_columns(header_row, fpath=fpath)
    entries = []
    maxcol = max(col.values())
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if row is None:
            continue
        padded = list(row) + [None] * max(0, maxcol + 1 - len(row))
        if len(padded) <= maxcol:
            continue
        amt = norm_amount(padded[col["umsatz"]])
        if amt is None:
            continue
        entries.append(make_entry(
            line=i,
            umsatz_raw=padded[col["umsatz"]],
            sh=padded[col["sh"]],
            konto=padded[col["konto"]],
            gegenkonto=padded[col["gegenkonto"]],
            datum=format_belegdatum(padded[col["datum"]]),
            belegfeld1=padded[col["belegfeld1"]],
            text=padded[col["buchungstext"]],
            amt=amt,
        ))
    if not entries:
        raise FormatError(
            f"'{fpath}' dosyasinda okunabilir hicbir veri satiri bulunamadi.\n"
            f"В файле '{fpath}' не найдено ни одной читаемой строки данных."
        )
    return entries


def load_datev_csv(fpath):
    rows = _read_csv_rows(fpath)
    if len(rows) < 2:
        raise FormatError(
            f"'{fpath}' dosyasi cok kisa veya bos.\n"
            f"Файл '{fpath}' слишком короткий или пустой."
        )
    header_idx = 1 if _header_has_datev(rows[1]) else (0 if _header_has_datev(rows[0]) else None)
    return _entries_from_datev_table(rows, fpath, header_idx)


def load_datev_xlsx(fpath):
    last_err = None
    for _title, rows in _xlsx_sheet_rows(fpath):
        for i, row in enumerate(rows[:30]):
            if row and _header_has_datev(row):
                try:
                    return _entries_from_datev_table(rows, fpath, i)
                except FormatError as ex:
                    # Bos header sheet: sonraki sheet'lere devam et.
                    last_err = ex
                    break
    if last_err is not None:
        raise last_err
    raise FormatError(
        f"'{fpath}' icinde DATEV baslik satiri bulunamadi.\n"
        f"В '{fpath}' не найдена строка заголовка DATEV."
    )


def _journal_colmap(header_row):
    col = {}
    for i, h in enumerate(header_row):
        name = _cell_str(h)
        if name:
            col[name] = i
    missing = [k for k in JOURNAL_REQUIRED if k not in col]
    if missing:
        raise FormatError(
            f"Journal dosyasinda eksik sutunlar: {missing}.\n"
            f"В Journal-файле отсутствуют столбцы: {missing}."
        )
    return col


def load_journal_xlsx(fpath):
    for _title, rows in _xlsx_sheet_rows(fpath):
        header_idx = None
        for i, row in enumerate(rows[:30]):
            if row and _header_has_journal(row):
                header_idx = i
                break
        if header_idx is None:
            continue
        col = _journal_colmap(rows[header_idx])
        entries = []
        for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not row or all(c is None or c == "" for c in row):
                continue
            need = max(col.values())
            padded = list(row) + [None] * (need + 1 - len(row))
            amt = norm_amount(padded[col["Betrag"]])
            # Keep 0,00 like DATEV (only skip unparseable amounts).
            if amt is None:
                continue
            soll = padded[col["Sollkto"]]
            haben = padded[col["Habenkto"]]
            if soll is None or haben is None or _cell_str(soll) == "" or _cell_str(haben) == "":
                continue
            entries.append(make_entry(
                line=i,
                umsatz_raw=None,
                sh="S",
                konto=soll,
                gegenkonto=haben,
                datum=format_belegdatum(padded[col["Belegdat."]]),
                belegfeld1=padded[col["Belegnr."]],
                text=padded[col["Buchungstext"]],
                amt=amt,
            ))
        if not entries:
            # Bos header sheet: sonraki sheet'lere devam et.
            continue
        return entries
    raise FormatError(
        f"'{fpath}' icinde Journal baslik satiri bulunamadi "
        f"(veya baslikli sheet'lerde okunabilir veri yok).\n"
        f"В '{fpath}' не найдена строка заголовка Journal "
        f"(или на листах с заголовком нет читаемых данных)."
    )


def load(fpath):
    """Dosyayi otomatik tanir ve ortak entry listesine cevirir."""
    fmt = detect_format(fpath)
    if fmt == FORMAT_DATEV_CSV:
        return load_datev_csv(fpath)
    if fmt == FORMAT_DATEV_XLSX:
        return load_datev_xlsx(fpath)
    if fmt == FORMAT_JOURNAL_XLSX:
        return load_journal_xlsx(fpath)
    raise FormatError(f"Bilinmeyen format: {fmt}")


def net_self_cancelling(entries):
    by_swap_key = defaultdict(list)
    for e in entries:
        by_swap_key[(e["amt"], e["sh"], e["konto"], e["gegenkonto"], e["datum"])].append(e)
    netted = 0
    for e in entries:
        if e["netted"] or not e["text"].startswith("<Storno>"):
            continue
        opp_key = (e["amt"], e["sh"], e["gegenkonto"], e["konto"], e["datum"])
        candidates = [c for c in by_swap_key.get(opp_key, []) if not c["netted"] and not c["text"].startswith("<Storno>")]
        if candidates:
            twin = candidates[0]
            e["netted"] = True
            twin["netted"] = True
            netted += 1
    return netted


def canon_sh_konto(e):
    """DATEV H kayitlarini Journal'in Soll bakisina cevirir (S + ters Konto).
    Boylece ayni ekonomik islem S/H yonu farkli export'larda eslesir."""
    if e["sh"] == "H":
        return "S", e["gegenkonto"]
    return e["sh"], e["konto"]


def key1(e):
    # Bos / bilimsel-notasyon Belegfeld ile eslestirme YAPMA - aksi halde
    # ayni tutar+hesap+tarihteki FARKLI kayitlar yanlis eslesir.
    if not e["beleg_n"]:
        return None
    sh, konto = canon_sh_konto(e)
    return (e["amt"], sh, konto, e["datum"], e["beleg_n"])


def key2(e):
    sh, konto = canon_sh_konto(e)
    return (e["amt"], sh, konto, e["datum"], e["text_n"])


def key3(e):
    sh, _konto = canon_sh_konto(e)
    return (e["amt"], sh, e["datum"], e["text_n"])


def key4(e):
    sh, konto = canon_sh_konto(e)
    return (e["amt"], sh, konto, e["datum"])


def key_value1(e):
    """Ayni kimlik (Belegfeld1+Konto+Tarih) ama Umsatz haric - deger farkini yakalar.
    Belegfeld1 BOS ise None doner (esnek eslesme YAPILMAZ) - aksi halde ayni
    Konto+Tarih'te Belegfeld1'i olmayan birden fazla FARKLI islem birbiriyle
    yanlislikla eslesir (Umsatz burada yok sayildigi icin tutar korumasi da
    devre disi kalir).
    Та же идентичность (Belegfeld1+Konto+Дата), но без Umsatz - ловит разницу
    в сумме. Если Belegfeld1 ПУСТОЙ, возвращает None (сопоставление НЕ
    делается) - иначе несколько РАЗНЫХ операций без Belegfeld1 в одном
    Konto+Дата ошибочно сопоставятся друг с другом (сумма здесь
    игнорируется, поэтому и защита по сумме отключена)."""
    if not e["beleg_n"]:
        return None
    sh, konto = canon_sh_konto(e)
    return (sh, konto, e["datum"], e["beleg_n"])


def key_value2(e):
    """Ayni kimlik (Metin+Konto+Tarih) ama Umsatz haric. Metin BOS ise None
    doner (esnek eslesme YAPILMAZ) - key_value1'deki ayni sebepten.
    Та же идентичность (Текст+Konto+Дата), но без Umsatz. Если текст ПУСТОЙ,
    возвращает None - по той же причине, что и key_value1."""
    if not e["text_n"]:
        return None
    sh, konto = canon_sh_konto(e)
    return (sh, konto, e["datum"], e["text_n"])


def match_tier(pool1, pool2, keyfunc, tier_name, keyfunc2=None, kind="MATCH", strict=False):
    """keyfunc2 verilmezse iki taraf da ayni fonksiyonla eslestirilir.
    Verilirse pool1 icin keyfunc, pool2 icin keyfunc2 kullanilir - bu, bir
    dosyada Belegfeld1/Buchungstext alanlarinin yer degistirdigi durumlari
    (STELLANTIS RE2026-0363 gibi) yakalamak icin gerekli.

    kind: "MATCH" (sari/yellow) veya "MISMATCH" (turuncu/orange, deger
    farkli/supheli). Bu deger dogrudan entry["pair_kind"] icine yazilir -
    birlesik sayfa (KARSILASTIRMA-SRAVNENIE) ve per-dosya sayfalari BUNA
    gore renklendirilir; tier ISMINE (string) bakarak tahmin ETMEZ - isim
    metnini degistirmek renklendirmeyi BOZMAZ.

    strict=True: bir anahtar icin HER IKI tarafta da TAM OLARAK 1 aday
    olmadikca eslestirme YAPILMAZ. Bu, "Zahlungseing.: diverse Rechnungen"
    gibi GENEL/tekrar eden metinlerle Umsatz'i yok sayan katmanlarda (deger
    farki, supheli) hayati onemde: aksi haldeayni genel metni paylasan
    ama gercekte FARKLI faturalara ait birden fazla kayit, liste sirasina
    gore RASTGELE birbirine eslenir (yanlis pozitif).

    Ortak uzun fatura/referans no'lari (ref_ids) VARSA ve kesisimleri bossa,
    bu aday cifti ATLENMEZ - ayni tutar+hesap+tarihteki farkli Vodafone
    faturalari gibi yanlis eslesmeleri engeller.
    """
    kf2 = keyfunc2 or keyfunc
    d1 = defaultdict(list)
    for e in pool1:
        k = keyfunc(e)
        if k is None:
            continue
        d1[k].append(e)
    d2 = defaultdict(list)
    for e in pool2:
        k = kf2(e)
        if k is None:
            continue
        d2[k].append(e)
    for k, l1 in d1.items():
        l2 = d2.get(k)
        if not l2:
            continue
        # ref_id catismasi olanlari ele
        pairs = []
        used2 = set()
        for a in l1:
            if a["matched"]:
                continue
            for b in l2:
                if b["matched"] or id(b) in used2:
                    continue
                ia = a.get("ref_ids") or set()
                ib = b.get("ref_ids") or set()
                if ia and ib and ia.isdisjoint(ib):
                    continue
                pairs.append((a, b))
                used2.add(id(b))
                break
        if strict:
            # strict: anahtarda her iki tarafta da filtreden sonra tek aday
            if len(l1) != 1 or len(l2) != 1:
                continue
            if len(pairs) != 1:
                continue
        for a, b in pairs:
            a["matched"] = True
            a["tier"] = tier_name
            a["pair"] = b
            a["pair_kind"] = kind
            b["matched"] = True
            b["tier"] = tier_name
            b["pair"] = a
            b["pair_kind"] = kind


def substring_id_tier(pool1, pool2, tier_name, min_len=5):
    """Son care eslesme: Umsatz+Soll/Haben+Belegdatum AYNI olmali (guclu
    guvenlik), Konto yok sayilir, kimlik ise bir tarafin Belegfeld1'inin
    (yeterince uzunsa) digerinin Buchungstext'i icinde GECMESI (substring)
    ile kontrol edilir. Bu, Konto yeniden kodlanirken Belegfeld1'in
    Buchungstext icine tasindigi durumlari yakalar (orn. BIV Group Trans
    'FS/26/01/2026' -> Buchungstext'te '...FS/26/01/2026' olarak gecmesi).
    Umsatz zaten esit oldugundan hep kind='MATCH' olarak isaretlenir.

    Финальный уровень: Umsatz+Soll/Haben+Belegdatum должны совпадать ТОЧНО
    (сильная защита), Konto игнорируется, идентичность проверяется как
    вхождение (substring) Belegfeld1 одной стороны (если достаточно длинный)
    в Buchungstext другой. Так ловятся случаи, когда при перекодировке Konto
    номер документа переехал в текст проводки. Так как Umsatz всегда равен,
    результат всегда kind='MATCH'."""
    by_amtdate = defaultdict(list)
    for e in pool2:
        sh, _ = canon_sh_konto(e)
        by_amtdate[(e["amt"], sh, e["datum"])].append(e)
    for e in pool1:
        if e["matched"]:
            continue
        sh, _ = canon_sh_konto(e)
        candidates = [c for c in by_amtdate.get((e["amt"], sh, e["datum"]), []) if not c["matched"]]
        if not candidates:
            continue
        beleg1 = e["belegfeld1"].strip().upper()
        hits = []
        for c in candidates:
            beleg2 = c["belegfeld1"].strip().upper()
            hit = (
                (len(beleg1) >= min_len and beleg1 in c["text"].upper())
                or (len(beleg2) >= min_len and beleg2 in e["text"].upper())
            )
            if not hit:
                continue
            ia = e.get("ref_ids") or set()
            ib = c.get("ref_ids") or set()
            if ia and ib and ia.isdisjoint(ib):
                continue
            hits.append(c)
        # Sadece TEK bir aday substring ile eslesiyorsa esle - birden fazla
        # aday varsa (belirsiz) HICBIRINI eslestirme (yanlis pozitif riski).
        # Сопоставляем, только если substring совпал РОВНО с одним
        # кандидатом - если их несколько (неоднозначно), не сопоставляем
        # ни с одним (риск ложного совпадения).
        if len(hits) == 1:
            c = hits[0]
            e["matched"] = True
            e["tier"] = tier_name
            e["pair"] = c
            e["pair_kind"] = "MATCH"
            c["matched"] = True
            c["tier"] = tier_name
            c["pair"] = e
            c["pair_kind"] = "MATCH"


def shared_ref_id_tier(pool1, pool2, tier_name, min_len=7):
    """Umsatz+SH+Belegdatum ayni iken Buchungstext/Belegfeld icindeki ortak
    uzun rakam kimligi (fatura no vb.) ile eslestir. Konto farkli olabilir
    (yeniden kodlama). Bilimsel-notasyon Belegfeld yuzunden key1'in
    kacirdigi / yanlis esledigi Vodafone tipi kayitlari duzeltir.
    Belirsiz (birden fazla aday) durumlarda eslestirmez."""
    by_amtdate = defaultdict(list)
    for e in pool2:
        sh, _ = canon_sh_konto(e)
        by_amtdate[(e["amt"], sh, e["datum"])].append(e)
    for e in pool1:
        if e["matched"]:
            continue
        ids = e.get("ref_ids") or extract_ref_ids(e["belegfeld1"], e["text"], min_len=min_len)
        if not ids:
            continue
        sh, _ = canon_sh_konto(e)
        hits = []
        for c in by_amtdate.get((e["amt"], sh, e["datum"]), []):
            if c["matched"]:
                continue
            ids2 = c.get("ref_ids") or extract_ref_ids(c["belegfeld1"], c["text"], min_len=min_len)
            if ids & ids2:
                hits.append(c)
        if len(hits) == 1:
            c = hits[0]
            e["matched"] = True
            e["tier"] = tier_name
            e["pair"] = c
            e["pair_kind"] = "MATCH"
            c["matched"] = True
            c["tier"] = tier_name
            c["pair"] = e
            c["pair_kind"] = "MATCH"


def _progress(progress_cb, percent, stage):
    if progress_cb is not None:
        progress_cb(int(percent), stage)


def compare(f1_path, f2_path, progress_cb=None):
    _progress(progress_cb, 5, "load1")
    e1 = load(f1_path)
    _progress(progress_cb, 15, "load2")
    e2 = load(f2_path)

    _progress(progress_cb, 25, "netting")
    net1 = net_self_cancelling(e1)
    net2 = net_self_cancelling(e2)
    for e in e1 + e2:
        if e["netted"]:
            e["matched"] = True
            e["tier"] = "IC-NETLESME (dosya icinde mukerrer+iptal, net sifir)"

    _progress(progress_cb, 35, "match_tier1")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key1, "TAM ESLESME / ПОЛНОЕ СОВПАДЕНИЕ (Belegfeld1+Konto+Tutar/Счёт+Сумма)", kind="MATCH", strict=True)

    _progress(progress_cb, 45, "match_tier2")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key2, "Konto+Tutar+Metin / Счёт+Сумма+Текст", kind="MATCH", strict=True)

    _progress(progress_cb, 52, "match_tier3")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key3, "Yeniden-kodlanmis / Перекодировано (Konto degisti / счёт изменён)", kind="MATCH", strict=True)

    # Yukaridaki 3 katman "strict" (1:1 belirsiz olmayan) calisti - ayni
    # anahtari paylasan ama BIRDEN FAZLA aday olan (yani hangisinin hangisiyle
    # eslestigi belirsiz olan) kayitlar kasten eslenmedi. Simdi bunlari
    # ikinci bir round'da DAHA GEVSEK (non-strict) calistiriyoruz: eger
    # tum adaylar birbirinin AYNISI ise (ayni Belegfeld1 VE ayni metin,
    # yani gercekten ayirt edilemez kopyalar) sirayla eslestirmek guvenlidir.
    # Verilerin GERCEKTEN farkli oldugu durumlarda (Sammelkunde vs Daimler
    # Truck AG gibi) zaten ayni anahtari paylasmazlar cunku bu ikinci tur
    # hala ayni key fonksiyonlarini kullanir - sadece "n aday var, hepsi
    # birbirinin klonu" durumunda devreye girer.
    # Верхние 3 уровня отработали в строгом режиме (без гарантированной
    # однозначности запись не сопоставлялась). Теперь второй проход, менее
    # строгий: если ВСЕ кандидаты идентичны друг другу (тот же Belegfeld1 И
    # тот же текст, то есть настоящие неотличимые дубликаты), сопоставление
    # по порядку безопасно.
    def _all_identical(entries):
        return len({(e["belegfeld1"], e["text"]) for e in entries}) == 1

    for kf, name in ((key1, "TAM ESLESME"), (key2, "Konto+Tutar+Metin"), (key3, "Yeniden-kodlanmis")):
        rem1 = [e for e in e1 if not e["matched"]]
        rem2 = [e for e in e2 if not e["matched"]]
        d1 = defaultdict(list)
        for e in rem1:
            k = kf(e)
            if k is not None:
                d1[k].append(e)
        d2 = defaultdict(list)
        for e in rem2:
            k = kf(e)
            if k is not None:
                d2[k].append(e)
        for k, l1 in d1.items():
            l2 = d2.get(k)
            if not l2:
                continue
            if not (_all_identical(l1) and _all_identical(l2)):
                continue
            n = min(len(l1), len(l2))
            for a, b in zip(l1[:n], l2[:n]):
                ia = a.get("ref_ids") or set()
                ib = b.get("ref_ids") or set()
                if ia and ib and ia.isdisjoint(ib):
                    continue
                a["matched"] = True
                a["tier"] = f"{name} (klon kayit / клон-запись)"
                a["pair"] = b
                a["pair_kind"] = "MATCH"
                b["matched"] = True
                b["tier"] = f"{name} (klon kayit / клон-запись)"
                b["pair"] = a
                b["pair_kind"] = "MATCH"

    # Konto yeniden kodlanirken Belegfeld1 (fatura/referans no) Buchungstext
    # icine tasinmis olabilir (orn. BIV Group Trans 'FS/26/01/2026'). Umsatz+
    # Datum+SH ayni olmasi sarti ile substring eslesmesi dene.
    # При перекодировке Konto номер документа (Belegfeld1) мог "переехать" в
    # текст проводки. Пробуем сопоставление по подстроке при условии, что
    # Umsatz+Datum+SH совпадают точно.
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    substring_id_tier(rem1, rem2, "Yeniden-kodlanmis / Перекодировано (Belegfeld1 metne tasinmis / номер переехал в текст)")

    # Ortak fatura/referans no (metin icindeki uzun rakamlar) + tutar+tarih
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    shared_ref_id_tier(
        rem1, rem2,
        "REF-ID ESLESME / СОВПАДЕНИЕ ПО НОМЕРУ (metindeki fatura/referans no)",
    )

    # DEGER-FARKI: ayni kimlik (Belegfeld1/Metin+Konto+Tarih) ama Umsatz farkli
    # -> "ayni kayit ama yanlis yazilmis" (STELLANTIS RE2026-0363 tipi hata)
    # ЗНАЧЕНИЕ ОТЛИЧАЕТСЯ: та же идентичность, но другая сумма
    # -> "та же запись, но написана неверно"
    _progress(progress_cb, 68, "match_value")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key_value1, "DEGER-FARKI / ЗНАЧЕНИЕ ОТЛИЧАЕТСЯ (Belegfeld1+Konto+Tarih ayni, Umsatz farkli)", kind="MISMATCH", strict=True)

    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key_value2, "DEGER-FARKI / ЗНАЧЕНИЕ ОТЛИЧАЕТСЯ (Metin+Konto+Tarih ayni, Umsatz farkli)", kind="MISMATCH", strict=True)

    # Capraz alan eslesmesi: bir dosyada Belegfeld1/Buchungstext yer
    # degistirmis olabilir (orn. birinde Belegfeld1="RE2026-0363", digerinde
    # Buchungstext="RE2026-0363" olabilir). Her iki yonde de dene.
    # Перекрёстное сопоставление полей: в одном файле Belegfeld1/Buchungstext
    # могут быть поменяны местами. Пробуем в обе стороны.
    _progress(progress_cb, 78, "match_cross")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key_value1, "DEGER-FARKI / ЗНАЧЕНИЕ ОТЛИЧАЕТСЯ (capraz alan / перекрёстное поле)", keyfunc2=key_value2, kind="MISMATCH", strict=True)

    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key_value2, "DEGER-FARKI / ЗНАЧЕНИЕ ОТЛИЧАЕТСЯ (capraz alan / перекрёстное поле)", keyfunc2=key_value1, kind="MISMATCH", strict=True)

    _progress(progress_cb, 85, "match_tier4")
    rem1 = [e for e in e1 if not e["matched"]]
    rem2 = [e for e in e2 if not e["matched"]]
    match_tier(rem1, rem2, key4, "SUPHELI-ESLESME / СОМНИТЕЛЬНОЕ (sadece tutar+konto+tarih / только сумма+счёт+дата)", kind="MISMATCH", strict=True)

    _progress(progress_cb, 90, "compare_done")
    return e1, e2, net1, net2


def account_balance_diff(e1, e2):
    def eb_wert(e):
        t = e["text"]
        return "EB-Wert" in t or "Gewinnvortrag" in t or "Zahllast" in t

    def balances(entries, skip_eb):
        bal = defaultdict(float)
        cnt = defaultdict(int)
        for e in entries:
            if skip_eb and eb_wert(e):
                continue
            sign = 1 if e["sh"] == "S" else -1
            bal[e["konto"]] += sign * e["amt"]
            cnt[e["konto"]] += 1
            bal[e["gegenkonto"]] -= sign * e["amt"]
            cnt[e["gegenkonto"]] += 1
        return bal, cnt

    # EB-Wert / Gewinnvortrag / Zahllast (acilis bakiyesi) satirlari her iki
    # dosyada da HER ZAMAN atlanir - aksi halde sadece bir dosyada acilis
    # bakiyesi varsa (ornegin biri sadece bir ay, digeri tum yil exportu ise)
    # hesap bakiyeleri yapay olarak cok farkli cikar.
    # Строки EB-Wert / Gewinnvortrag / Zahllast (входящий остаток) ВСЕГДА
    # исключаются в обоих файлах - иначе баланс счёта будет искусственно
    # отличаться, если входящий остаток есть только в одном файле.
    bal1, cnt1 = balances(e1, skip_eb=True)
    bal2, cnt2 = balances(e2, skip_eb=True)

    all_k = set(bal1) | set(bal2)
    diffs = []
    for k in all_k:
        b1 = round(bal1.get(k, 0.0), 2)
        b2 = round(bal2.get(k, 0.0), 2)
        d = round(b2 - b1, 2)
        if abs(d) > 0.01:
            diffs.append((k, b1, b2, d, cnt1.get(k, 0), cnt2.get(k, 0)))
    diffs.sort(key=lambda x: -abs(x[3]))
    return diffs


def date_sort_key(datum):
    """Belegdatum 'TTAA' (gun+ay) formatinda - kronolojik siralama icin (ay, gun) dondurur.
    Belegdatum в формате 'ДДММ' (день+месяц) - возвращает (месяц, день) для хронологической сортировки."""
    d = (datum or "").strip()
    if len(d) == 4 and d.isdigit():
        return (int(d[2:4]), int(d[0:2]))
    return (99, 99)


def build_combined_rows(e1, e2):
    """Iki dosyayi tek bir listede 'grup' halinde esler: her grup 1-2 satirdan
    olusur (MATCH: her iki taraf ayni: sari / MISMATCH: ayni kayit ama bir
    alani farkli: turuncu / ONLY1-ONLY2: sadece bir tarafta var, digeri bos+
    kirmizi). Sonuc kronolojik olarak siralanir.

    Объединяет оба файла в один список 'групп': каждая группа состоит из 1-2
    строк (MATCH: одинаково в обоих - жёлтый / MISMATCH: та же запись, но
    поле отличается - оранжевый / ONLY1-ONLY2: есть только на одной стороне,
    другая пустая+красная). Результат сортируется хронологически."""
    groups = []
    consumed_e2_ids = set()

    for e in e1:
        if e["netted"]:
            continue
        if e["pair"] is not None:
            partner = e["pair"]
            kind = e["pair_kind"] or "MATCH"
            groups.append({"kind": kind, "e1": e, "e2": partner})
            consumed_e2_ids.add(id(partner))
        else:
            groups.append({"kind": "ONLY1", "e1": e, "e2": None})

    for e in e2:
        if e["netted"] or id(e) in consumed_e2_ids:
            continue
        if e["pair"] is None:
            groups.append({"kind": "ONLY2", "e1": None, "e2": e})
        # e["pair"] is not None ama consumed_e2_ids'te yoksa (teorik olarak
        # olmamali, ama guvenlik icin): zaten e1 tarafindan eslenmis kabul edilir.

    def sort_key(g):
        ref = g["e1"] or g["e2"]
        return (date_sort_key(ref["datum"]), ref["konto"], -abs(ref["amt"]))

    groups.sort(key=sort_key)
    return groups


def diff_note(e1, e2):
    if e1 is None or e2 is None:
        return ""
    fields = [
        ("Umsatz/Сумма", "umsatz"),
        ("Konto/Счёт", "konto"),
        ("Gegenkonto/Контрсчёт", "gegenkonto"),
        ("Belegfeld1/Номер", "belegfeld1"),
        ("Buchungstext/Текст", "text"),
    ]
    diffs = []
    for label, key in fields:
        if e1[key] != e2[key]:
            diffs.append(f"{label}: '{e1[key]}' -> '{e2[key]}'")
    return " | ".join(diffs)


# ---------------- Excel raporu ----------------

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
GREEN = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)

INVALID_SHEET_CHARS = '[]:*?/\\'


def _norm_report_lang(lang):
    import i18n
    lang = (lang or "tr").lower()
    return lang if lang in i18n.SUPPORTED else "tr"


def _excel_t(key, lang, **kwargs):
    import i18n
    return i18n.t(key, lang=_norm_report_lang(lang), **kwargs)


def excel_entry_headers(lang=None):
    return [
        _excel_t("excel_h_line", lang),
        _excel_t("excel_h_amount", lang),
        _excel_t("excel_h_sh", lang),
        _excel_t("excel_h_konto", lang),
        _excel_t("excel_h_gegen", lang),
        _excel_t("excel_h_date", lang),
        _excel_t("excel_h_beleg", lang),
        _excel_t("excel_h_text", lang),
        _excel_t("excel_h_status", lang),
    ]


def excel_combined_headers(lang=None):
    return [
        _excel_t("excel_h_source", lang),
        _excel_t("excel_h_line", lang),
        _excel_t("excel_h_amount", lang),
        _excel_t("excel_h_sh", lang),
        _excel_t("excel_h_konto", lang),
        _excel_t("excel_h_gegen", lang),
        _excel_t("excel_h_date", lang),
        _excel_t("excel_h_beleg", lang),
        _excel_t("excel_h_text", lang),
        _excel_t("excel_h_note", lang),
        _excel_t("excel_h_status", lang),
    ]


# Default TR bilingual headers (tests / callers without lang).
HEADERS = [
    "Satir / Строка",
    "Umsatz / Сумма",
    "Soll/Haben / Дебет-Кредит",
    "Konto / Счёт",
    "Gegenkonto / Контрсчёт",
    "Belegdatum / Дата документа",
    "Belegfeld 1 / Номер документа",
    "Buchungstext / Текст проводки",
    "Eslesme Durumu / Статус сопоставления",
]


def sanitize_sheet_name(name, fallback, taken):
    """Excel sayfa adi kurallarina uydurur (gecersiz karakterleri temizler,
    31 karaktere kisaltir, bos/cakisan isimleri benzersiz yapar).
    Приводит имя к правилам листа Excel (убирает недопустимые символы,
    обрезает до 31 символа, делает уникальным при совпадении)."""
    cleaned = "".join(c for c in (name or "") if c not in INVALID_SHEET_CHARS).strip()
    if not cleaned:
        cleaned = fallback
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 1
    while cleaned.lower() in taken:
        suffix_str = f"_{suffix}"
        cleaned = base[: 31 - len(suffix_str)] + suffix_str
        suffix += 1
    taken.add(cleaned.lower())
    return cleaned


def fill_for(e):
    if not e["matched"]:
        return RED
    if e["tier"] and e["tier"].startswith("IC-NETLESME"):
        return GREEN
    if e["pair_kind"] == "MISMATCH":
        return ORANGE
    return YELLOW


def write_header(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def write_entry_sheet(ws, entries, lang=None):
    headers = excel_entry_headers(lang)
    write_header(ws, headers)
    only_status = _excel_t("excel_only_here", lang)
    for e in entries:
        status = e["tier"] if e["matched"] else only_status
        ws.append([e["line"], e["umsatz"], e["sh"], e["konto"], e["gegenkonto"], e["datum"], e["belegfeld1"], e["text"], status])
        r = ws.max_row
        fill = fill_for(e)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
    widths = [7, 12, 10, 9, 11, 11, 26, 55, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


COMBINED_HEADERS = [
    "Kaynak / Источник",
    "Satir / Строка",
    "Umsatz / Сумма",
    "Soll/Haben / Дебет-Кредит",
    "Konto / Счёт",
    "Gegenkonto / Контрсчёт",
    "Belegdatum / Дата документа",
    "Belegfeld 1 / Номер документа",
    "Buchungstext / Текст проводки",
    "Not / Примечание",
    "Durum / Status",
]

THIN_TOP = Border(top=Side(style="thin", color="808080"))


def write_combined_sheet(ws, groups, f1_label, f2_label, lang=None):
    headers = excel_combined_headers(lang)
    write_header(ws, headers)
    missing = _excel_t("excel_missing", lang)
    row = 2

    def emit(entry, source_label, note, fill, top_border, status):
        nonlocal row
        if entry is None:
            values = [source_label, "", "", "", "", "", "", "", "", note, status]
        else:
            values = [source_label, entry["line"], entry["umsatz"], entry["sh"], entry["konto"],
                      entry["gegenkonto"], entry["datum"], entry["belegfeld1"], entry["text"], note, status]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill
            if top_border:
                cell.border = THIN_TOP
        row += 1

    for g in groups:
        kind = g["kind"]
        if kind == "MATCH":
            emit(g["e1"], f1_label, "", YELLOW, True, "MATCH")
            emit(g["e2"], f2_label, "", YELLOW, False, "MATCH")
        elif kind == "MISMATCH":
            note = diff_note(g["e1"], g["e2"])
            emit(g["e1"], f1_label, "", ORANGE, True, "MISMATCH")
            emit(g["e2"], f2_label, note, ORANGE, False, "MISMATCH")
        elif kind == "ONLY1":
            # Both half-rows RED: present row = only in this file; blank = missing opposite.
            emit(g["e1"], f1_label, "", RED, True, "ONLY1")
            emit(None, f2_label, missing, RED, False, "ONLY1")
        elif kind == "ONLY2":
            emit(None, f1_label, missing, RED, True, "ONLY2")
            emit(g["e2"], f2_label, "", RED, False, "ONLY2")

    widths = [16, 7, 12, 10, 9, 11, 11, 26, 50, 45, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:K{row - 1}"


def write_filter_sheet(ws, groups, f1_label, f2_label, kinds, lang=None):
    """Combined-sheet satirlari arasindan sadece secili Durum turlerini yazar."""
    headers = excel_combined_headers(lang)
    write_header(ws, headers)
    missing = _excel_t("excel_missing", lang)
    row = 2

    def emit(entry, source_label, note, fill, top_border, status):
        nonlocal row
        if entry is None:
            values = [source_label, "", "", "", "", "", "", "", "", note, status]
        else:
            values = [source_label, entry["line"], entry["umsatz"], entry["sh"], entry["konto"],
                      entry["gegenkonto"], entry["datum"], entry["belegfeld1"], entry["text"], note, status]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill
            if top_border:
                cell.border = THIN_TOP
        row += 1

    for g in groups:
        kind = g["kind"]
        if kind not in kinds:
            continue
        if kind == "MISMATCH":
            note = diff_note(g["e1"], g["e2"])
            emit(g["e1"], f1_label, "", ORANGE, True, "MISMATCH")
            emit(g["e2"], f2_label, note, ORANGE, False, "MISMATCH")
        elif kind == "ONLY1":
            emit(g["e1"], f1_label, "", RED, True, "ONLY1")
            emit(None, f2_label, missing, RED, False, "ONLY1")
        elif kind == "ONLY2":
            emit(None, f1_label, missing, RED, True, "ONLY2")
            emit(g["e2"], f2_label, "", RED, False, "ONLY2")

    widths = [16, 7, 12, 10, 9, 11, 11, 26, 50, 45, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:K{row - 1}"


def build_report(f1_path, f2_path, out_path, f1_label="Dosya1", f2_label="Dosya2",
                 progress_cb=None, write_html=False, write_pdf=False, detail_sheets=True,
                 lang=None):
    lang = _norm_report_lang(lang)
    e1, e2, net1, net2 = compare(f1_path, f2_path, progress_cb=progress_cb)

    _progress(progress_cb, 92, "build_sheets")
    wb = openpyxl.Workbook()

    ws_combined = wb.active
    ws_combined.title = "KARSILASTIRMA-SRAVNENIE"
    groups = build_combined_rows(e1, e2)
    write_combined_sheet(ws_combined, groups, f1_label, f2_label, lang=lang)

    ws_red = wb.create_sheet("FILTRE-KIRMIZI")
    write_filter_sheet(ws_red, groups, f1_label, f2_label, {"ONLY1", "ONLY2"}, lang=lang)
    ws_orange = wb.create_sheet("FILTRE-TURUNCU")
    write_filter_sheet(ws_orange, groups, f1_label, f2_label, {"MISMATCH"}, lang=lang)

    ws0 = wb.create_sheet("OZET-SVODKA")

    # NOT: ws.append([]) openpyxl'de max_row'u guncellemiyor (bos satirlarda
    # geriye kalabiliyor), bu yuzden burada satir numarasini KENDIMIZ manuel
    # sayiyoruz - ws.append()/ws.max_row'a guvenmiyoruz.
    row = 1

    def put(*values):
        nonlocal row
        for c, v in enumerate(values, start=1):
            ws0.cell(row=row, column=c, value=v)
        r = row
        row += 1
        return r

    r = put(_excel_t("excel_summary_title", lang))
    ws0.cell(row=r, column=1).font = Font(bold=True, size=14)
    put()
    put(_excel_t("excel_file1", lang), f1_path)
    put(_excel_t("excel_file2", lang), f2_path)
    put()

    matched1 = sum(1 for e in e1 if e["matched"])
    matched2 = sum(1 for e in e2 if e["matched"])
    red1 = len(e1) - matched1
    red2 = len(e2) - matched2

    r = put(
        "",
        _excel_t("excel_col_total", lang),
        _excel_t("excel_col_matched", lang),
        _excel_t("excel_col_only", lang),
    )
    for c in range(1, 5):
        ws0.cell(row=r, column=c).font = BOLD
    put(f1_label, len(e1), matched1, red1)
    put(f2_label, len(e2), matched2, red2)
    put()
    put(_excel_t("excel_net_dup", lang), f1_label, net1)
    put(_excel_t("excel_net_dup", lang), f2_label, net2)
    put()

    match_count = sum(1 for g in groups if g["kind"] == "MATCH")
    mismatch_count = sum(1 for g in groups if g["kind"] == "MISMATCH")
    only1_count = sum(1 for g in groups if g["kind"] == "ONLY1")
    only2_count = sum(1 for g in groups if g["kind"] == "ONLY2")
    put(_excel_t("excel_combined_summary", lang))
    put(_excel_t("excel_same_yellow", lang), match_count)
    put(_excel_t("excel_diff_orange", lang), mismatch_count)
    put(_excel_t("excel_only_f1", lang, name=f1_label), only1_count)
    put(_excel_t("excel_only_f2", lang, name=f2_label), only2_count)
    put()
    put(_excel_t("excel_legend_title", lang))

    legend_fills = [
        (_excel_t("excel_legend_yellow", lang), _excel_t("excel_legend_yellow_desc", lang), YELLOW),
        (_excel_t("excel_legend_orange", lang), _excel_t("excel_legend_orange_desc", lang), ORANGE),
        (_excel_t("excel_legend_green", lang), _excel_t("excel_legend_green_desc", lang), GREEN),
        (_excel_t("excel_legend_red", lang), _excel_t("excel_legend_red_desc", lang), RED),
    ]
    for label, desc, fill in legend_fills:
        r = put(label, desc)
        ws0.cell(row=r, column=1).font = BOLD
        ws0.cell(row=r, column=1).fill = fill

    for col, w in zip("ABCD", [55, 20, 40, 40]):
        ws0.column_dimensions[col].width = w

    diffs = account_balance_diff(e1, e2)
    entry_headers = excel_entry_headers(lang)
    bal_headers = [
        _excel_t("excel_h_konto", lang),
        _excel_t("excel_bal", lang, name=f1_label),
        _excel_t("excel_bal", lang, name=f2_label),
        _excel_t("excel_diff", lang),
        _excel_t("excel_n_entries", lang, name=f1_label),
        _excel_t("excel_n_entries", lang, name=f2_label),
    ]

    if detail_sheets:
        taken_sheet_names = {
            "karsilastirma-sravnenie", "filtre-kirmizi", "filtre-turuncu",
            "ozet-svodka", "farklar-raznica", "hesap farki-scheta",
        }
        sheet1_name = sanitize_sheet_name(f1_label, "Dosya1-Fayl1", taken_sheet_names)
        sheet2_name = sanitize_sheet_name(f2_label, "Dosya2-Fayl2", taken_sheet_names)

        ws1 = wb.create_sheet(sheet1_name)
        write_entry_sheet(ws1, e1, lang=lang)
        ws2 = wb.create_sheet(sheet2_name)
        write_entry_sheet(ws2, e2, lang=lang)

        ws3 = wb.create_sheet("FARKLAR-RAZNICA")
        write_header(ws3, [_excel_t("excel_source_file", lang)] + entry_headers[:8])
        for label, entries in ((f1_label, e1), (f2_label, e2)):
            for e in entries:
                if not e["matched"]:
                    ws3.append([
                        label, e["line"], e["umsatz"], e["sh"], e["konto"],
                        e["gegenkonto"], e["datum"], e["belegfeld1"], e["text"],
                    ])
                    for c in range(1, 10):
                        ws3.cell(row=ws3.max_row, column=c).fill = RED
        widths3 = [18, 7, 12, 10, 9, 11, 11, 26, 55]
        for i, w in enumerate(widths3, start=1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

        ws4 = wb.create_sheet("HESAP FARKI-SCHETA")
        write_header(ws4, bal_headers)
        for k, b1, b2, d, c1, c2 in diffs:
            ws4.append([k, b1, b2, d, c1, c2])
            if abs(d) >= 1000:
                for c in range(1, 7):
                    ws4.cell(row=ws4.max_row, column=c).fill = RED
        for i, w in enumerate([10, 16, 16, 14, 12, 12], start=1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws4.freeze_panes = "A2"
    else:
        ws4 = wb.create_sheet("HESAP FARKI-SCHETA")
        write_header(ws4, bal_headers)
        for k, b1, b2, d, c1, c2 in diffs:
            ws4.append([k, b1, b2, d, c1, c2])
        for i, w in enumerate([10, 16, 16, 14, 12, 12], start=1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws4.freeze_panes = "A2"

    wb.save(out_path)

    html_path = pdf_path = None
    warnings = []
    base, _ = os.path.splitext(out_path)
    if write_html or write_pdf:
        from report_extra import write_html_report, write_pdf_report
        common = dict(
            f1_path=f1_path, f2_path=f2_path, f1_label=f1_label, f2_label=f2_label,
            match_count=match_count, mismatch_count=mismatch_count,
            only1_count=only1_count, only2_count=only2_count,
            groups=groups, diffs=diffs, lang=lang,
        )
        if write_html:
            try:
                html_path = write_html_report(base + ".html", **common)
            except Exception as ex:
                warnings.append(f"html: {ex}")
        if write_pdf:
            try:
                pdf_path = write_pdf_report(base + ".pdf", **common)
            except Exception as ex:
                warnings.append(f"pdf: {ex}")

    _progress(progress_cb, 100, "done")
    return {
        "f1_total": len(e1), "f2_total": len(e2),
        "f1_red": red1, "f2_red": red2,
        "net1": net1, "net2": net2,
        "diffs": diffs,
        "out_path": out_path,
        "html_path": html_path,
        "pdf_path": pdf_path,
        "match_count": match_count,
        "mismatch_count": mismatch_count,
        "only1_count": only1_count,
        "only2_count": only2_count,
        "warnings": warnings,
    }
